from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from datetime import datetime, timedelta
from decimal import Decimal
import json
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Simulation, Supplier, Cargo, Customer, Plant, PlantInventory, APIConfiguration
from .forms import (
    SimulationForm, SupplierForm, CargoForm, CustomerForm, PlantForm, PlantInventoryForm,
    APIConfigurationForm, JSONUploadForm, MasterSimulationForm
)


@login_required
def dashboard(request):
    """Main dashboard view"""
    # Check for master simulation first
    master_simulation = Simulation.objects.filter(is_master=True).first()
    
    # Get user's simulations
    simulations = Simulation.objects.filter(user=request.user, is_master=False)
    active_simulation = simulations.filter(is_active=True).first()
    
    # If no active simulation, show master as default (read-only)
    if not active_simulation and master_simulation:
        active_simulation = master_simulation
    elif not active_simulation and simulations.exists():
        active_simulation = simulations.first()
        active_simulation.is_active = True
        active_simulation.save()
    
    context = {
        'simulations': simulations,
        'active_simulation': active_simulation,
        'master_simulation': master_simulation,  # Make sure this is passed
        'is_viewing_master': active_simulation and active_simulation.is_master if active_simulation else False,
    }
    
    if active_simulation:
        context.update(get_simulation_data(active_simulation))
    
    print(f"DEBUG - User: {request.user.username}, is_staff: {request.user.is_staff}")
    print(f"DEBUG - Master exists: {master_simulation is not None}")
    print(f"DEBUG - Master: {master_simulation}")
    
    return render(request, 'lng_planner/dashboard.html', context)


@login_required
def create_simulation(request):
    """Create a new simulation by copying from master"""
    master = Simulation.objects.filter(is_master=True).first()
    
    if not master:
        messages.error(request, 'No master simulation found. Please contact administrator.')
        return redirect('lng_planner:dashboard')
    
    if request.method == 'POST':
        # Get simulation name from form
        sim_name = request.POST.get('name', f"My Simulation - {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        
        # Create new simulation for user (copy from master)
        new_sim = Simulation.objects.create(
            user=request.user,
            name=sim_name,
            start_date=master.start_date,
            end_date=master.end_date,
            is_master=False,
            is_active=True
        )
        
        # Deactivate other user simulations
        Simulation.objects.filter(user=request.user).exclude(pk=new_sim.pk).update(is_active=False)
        
        # Copy plant inventories from master
        for pi in master.plant_inventories.all():
            PlantInventory.objects.create(
                simulation=new_sim,
                plant=pi.plant,
                opening_inventory=pi.opening_inventory
            )
        
        # Copy suppliers from master
        for supplier in master.suppliers.all():
            Supplier.objects.create(
                simulation=new_sim,
                plant=supplier.plant,
                name=supplier.name,
                daily_supply=supplier.daily_supply,
                from_date=supplier.from_date,
                to_date=supplier.to_date
            )
        
        # Copy cargos from master
        for cargo in master.cargos.all():
            Cargo.objects.create(
                simulation=new_sim,
                plant=cargo.plant,
                cargo_name=cargo.cargo_name,
                delivery_date=cargo.delivery_date,
                amount=cargo.amount
            )
        
        # Copy customers from master
        for customer in master.customers.all():
            Customer.objects.create(
                simulation=new_sim,
                plant=customer.plant,
                name=customer.name,
                daily_demand=customer.daily_demand,
                from_date=customer.from_date,
                to_date=customer.to_date
            )
        
        messages.success(request, f'Simulation "{new_sim.name}" created from master! You can now edit it.')
        return redirect('lng_planner:dashboard')
    
    # GET request - show simple form
    context = {
        'master': master,
        'suggested_name': f"My Simulation - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    }
    return render(request, 'lng_planner/create_from_master.html', context)


@login_required
def switch_simulation(request, simulation_id):
    """Switch active simulation"""
    simulation = get_object_or_404(Simulation, pk=simulation_id)
    
    # Allow switching to master or own simulations
    if not simulation.is_master and simulation.user != request.user:
        messages.error(request, 'You cannot access this simulation')
        return redirect('lng_planner:dashboard')
    
    Simulation.objects.filter(user=request.user).update(is_active=False)
    
    if not simulation.is_master:
        simulation.is_active = True
        simulation.save()
    
    messages.success(request, f'Switched to simulation: {simulation.name}')
    return redirect('lng_planner:dashboard')


@login_required
def delete_simulation(request, simulation_id):
    """Delete a simulation"""
    simulation = get_object_or_404(Simulation, pk=simulation_id, user=request.user)
    simulation.delete()
    messages.success(request, 'Simulation deleted successfully!')
    return redirect('lng_planner:dashboard')


# PLANT MANAGEMENT
@login_required
def manage_plants(request):
    """Manage plants"""
    plants = Plant.objects.all()
    
    if request.method == 'POST':
        form = PlantForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Plant added successfully!')
            return redirect('lng_planner:manage_plants')
    else:
        form = PlantForm()
    
    return render(request, 'lng_planner/manage_plants.html', {
        'plants': plants,
        'form': form
    })


@login_required
def delete_plant(request, plant_id):
    """Delete a plant"""
    plant = get_object_or_404(Plant, pk=plant_id)
    plant.delete()
    messages.success(request, 'Plant deleted successfully!')
    return redirect('lng_planner:manage_plants')


@login_required
def setup_plants(request, simulation_id):
    """Setup plant inventories for simulation"""
    simulation = get_object_or_404(Simulation, pk=simulation_id, user=request.user)
    plants = Plant.objects.all()
    
    if not plants.exists():
        messages.warning(request, 'Please create plants first.')
        return redirect('lng_planner:manage_plants')
    
    if request.method == 'POST':
        # Process plant inventories
        for plant in plants:
            inventory = request.POST.get(f'plant_{plant.id}')
            if inventory:
                PlantInventory.objects.update_or_create(
                    simulation=simulation,
                    plant=plant,
                    defaults={'opening_inventory': float(inventory)}
                )
        
        # Initialize with sample data
        initialize_sample_data(simulation)
        
        messages.success(request, 'Plant inventories configured successfully!')
        return redirect('lng_planner:dashboard')
    
    # Get existing plant inventories
    existing_inventories = {
        pi.plant_id: pi.opening_inventory 
        for pi in simulation.plant_inventories.all()
    }
    
    return render(request, 'lng_planner/setup_plants.html', {
        'simulation': simulation,
        'plants': plants,
        'existing_inventories': existing_inventories
    })


# SUPPLIER VIEWS
@login_required
def add_supplier(request, simulation_id):
    """Add new supplier"""
    simulation = get_object_or_404(Simulation, pk=simulation_id, user=request.user)
    
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            supplier = form.save(commit=False)
            supplier.simulation = simulation
            supplier.save()
            messages.success(request, 'Supplier added successfully!')
            return redirect('lng_planner:dashboard')
    else:
        form = SupplierForm(initial={
            'from_date': simulation.start_date,
            'to_date': simulation.end_date
        })
    
    return render(request, 'lng_planner/supplier_form.html', {
        'form': form,
        'simulation': simulation,
        'title': 'Add Supplier'
    })


@login_required
def edit_supplier(request, supplier_id):
    """Edit existing supplier"""
    supplier = get_object_or_404(Supplier, pk=supplier_id, simulation__user=request.user)
    simulation = supplier.simulation
    
    if request.method == 'POST':
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            messages.success(request, 'Supplier updated successfully!')
            return redirect('lng_planner:dashboard')
    else:
        form = SupplierForm(instance=supplier)
    
    return render(request, 'lng_planner/supplier_form.html', {
        'form': form,
        'simulation': simulation,
        'title': 'Edit Supplier'
    })


@login_required
def delete_supplier(request, supplier_id):
    """Delete a supplier"""
    supplier = get_object_or_404(Supplier, pk=supplier_id, simulation__user=request.user)
    supplier.delete()
    messages.success(request, 'Supplier deleted successfully!')
    return redirect('lng_planner:dashboard')


# CARGO VIEWS
@login_required
def add_cargo(request, simulation_id):
    """Add new cargo"""
    simulation = get_object_or_404(Simulation, pk=simulation_id, user=request.user)
    
    if request.method == 'POST':
        form = CargoForm(request.POST)
        if form.is_valid():
            cargo = form.save(commit=False)
            cargo.simulation = simulation
            cargo.save()
            messages.success(request, 'Cargo added successfully!')
            return redirect('lng_planner:dashboard')
    else:
        form = CargoForm()
    
    return render(request, 'lng_planner/cargo_form.html', {
        'form': form,
        'simulation': simulation,
        'title': 'Add Cargo'
    })


@login_required
def edit_cargo(request, cargo_id):
    """Edit existing cargo"""
    cargo = get_object_or_404(Cargo, pk=cargo_id, simulation__user=request.user)
    simulation = cargo.simulation
    
    if request.method == 'POST':
        form = CargoForm(request.POST, instance=cargo)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cargo updated successfully!')
            return redirect('lng_planner:dashboard')
    else:
        form = CargoForm(instance=cargo)
    
    return render(request, 'lng_planner/cargo_form.html', {
        'form': form,
        'simulation': simulation,
        'title': 'Edit Cargo'
    })


@login_required
def delete_cargo(request, cargo_id):
    """Delete a cargo"""
    cargo = get_object_or_404(Cargo, pk=cargo_id, simulation__user=request.user)
    cargo.delete()
    messages.success(request, 'Cargo deleted successfully!')
    return redirect('lng_planner:dashboard')


# CUSTOMER VIEWS
@login_required
def add_customer(request, simulation_id):
    """Add new customer"""
    simulation = get_object_or_404(Simulation, pk=simulation_id, user=request.user)
    
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            customer = form.save(commit=False)
            customer.simulation = simulation
            customer.save()
            messages.success(request, 'Customer added successfully!')
            return redirect('lng_planner:dashboard')
    else:
        form = CustomerForm(initial={
            'from_date': simulation.start_date,
            'to_date': simulation.end_date
        })
    
    return render(request, 'lng_planner/customer_form.html', {
        'form': form,
        'simulation': simulation,
        'title': 'Add Customer'
    })


@login_required
def edit_customer(request, customer_id):
    """Edit existing customer"""
    customer = get_object_or_404(Customer, pk=customer_id, simulation__user=request.user)
    simulation = customer.simulation
    
    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            messages.success(request, 'Customer updated successfully!')
            return redirect('lng_planner:dashboard')
    else:
        form = CustomerForm(instance=customer)
    
    return render(request, 'lng_planner/customer_form.html', {
        'form': form,
        'simulation': simulation,
        'title': 'Edit Customer'
    })


@login_required
def delete_customer(request, customer_id):
    """Delete a customer"""
    customer = get_object_or_404(Customer, pk=customer_id, simulation__user=request.user)
    customer.delete()
    messages.success(request, 'Customer deleted successfully!')
    return redirect('lng_planner:dashboard')


# EXPORT/IMPORT VIEWS
@login_required
def export_excel(request, simulation_id):
    """Export simulation data to Excel with plant-wise breakdown"""
    simulation = get_object_or_404(Simulation, pk=simulation_id, user=request.user)
    daily_data = calculate_daily_data(simulation)
    plants = Plant.objects.filter(
        id__in=simulation.plant_inventories.values_list('plant_id', flat=True)
    )
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'LNG Planning'
    
    # Header row with dates
    header = ['Item']
    for day in daily_data:
        header.append(day['date'].strftime('%b %d, %Y'))
    ws.append(header)
    
    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # For each plant
    for plant in plants:
        # Plant header row
        plant_header_row = [f'=== {plant.name} ===']
        plant_header_row.extend(['' for _ in daily_data])
        ws.append(plant_header_row)
        
        # Suppliers for this plant
        for supplier in simulation.suppliers.filter(plant=plant):
            row = [f'  {supplier.name}']
            for day in daily_data:
                plant_data = day['plant_data'].get(plant.id, {})
                supply = next((s['amount'] for s in plant_data.get('supplies', []) 
                             if s['type'] == 'supplier' and s['name'] == supplier.name), None)
                row.append(supply if supply else '-')
            ws.append(row)
        
        # Cargos for this plant
        for cargo in simulation.cargos.filter(plant=plant):
            row = [f'  {cargo.cargo_name}']
            for day in daily_data:
                plant_data = day['plant_data'].get(plant.id, {})
                supply = next((s['amount'] for s in plant_data.get('supplies', []) 
                             if s['type'] == 'cargo' and s['name'] == cargo.cargo_name), None)
                row.append(supply if supply else '-')
            ws.append(row)
        
        # Plant Total Supply
        row = [f'  Total Supply - {plant.name}']
        for day in daily_data:
            plant_data = day['plant_data'].get(plant.id, {})
            supply = plant_data.get('supply', 0)
            row.append(supply if supply > 0 else '-')
        ws.append(row)
        
        # Customers for this plant
        for customer in simulation.customers.filter(plant=plant):
            row = [f'  {customer.name}']
            for day in daily_data:
                plant_data = day['plant_data'].get(plant.id, {})
                demand = next((d['amount'] for d in plant_data.get('demands', []) 
                             if d['customer'] == customer.name), None)
                row.append(demand if demand else '-')
            ws.append(row)
        
        # Plant Total Demand
        row = [f'  Total Demand - {plant.name}']
        for day in daily_data:
            plant_data = day['plant_data'].get(plant.id, {})
            demand = plant_data.get('demand', 0)
            row.append(demand if demand > 0 else '-')
        ws.append(row)
        
        # Plant Closing Inventory
        row = [f'  Closing Inventory - {plant.name}']
        for day in daily_data:
            plant_data = day['plant_data'].get(plant.id, {})
            inventory = plant_data.get('inventory', 0)
            row.append(inventory)
        ws.append(row)
        
        # Add blank row between plants
        ws.append([''])
    
    # Overall Totals Section
    ws.append(['=== TOTAL (All Plants) ==='])
    
    # Total Supply (all plants)
    row = ['Total Supply (All Plants)']
    for day in daily_data:
        row.append(day['total_supply'] if day['total_supply'] > 0 else '-')
    ws.append(row)
    
    # Total Demand (all plants)
    row = ['Total Demand (All Plants)']
    for day in daily_data:
        row.append(day['total_demand'] if day['total_demand'] > 0 else '-')
    ws.append(row)
    
    # Total Closing Inventory (all plants)
    row = ['Total Closing Inventory (All Plants)']
    for day in daily_data:
        row.append(day['total_inventory'])
    ws.append(row)
    
    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=LNG_Planning_{simulation.start_date}_to_{simulation.end_date}.xlsx'
    wb.save(response)
    
    return response


@login_required
def export_json(request, simulation_id):
    """Export simulation as JSON"""
    simulation = get_object_or_404(Simulation, pk=simulation_id, user=request.user)
    
    data = {
        'simulation': {
            'name': simulation.name,
            'start_date': str(simulation.start_date),
            'end_date': str(simulation.end_date),
            'opening_inventory': float(simulation.opening_inventory)
        },
        'suppliers': [
            {
                'name': s.name,
                'daily_supply': float(s.daily_supply),
                'from_date': str(s.from_date),
                'to_date': str(s.to_date)
            }
            for s in simulation.suppliers.all()
        ],
        'cargos': [
            {
                'cargo_name': c.cargo_name,
                'date': str(c.delivery_date),
                'amount': float(c.amount)
            }
            for c in simulation.cargos.all()
        ],
        'customers': [
            {
                'name': c.name,
                'daily_demand': float(c.daily_demand),
                'from_date': str(c.from_date),
                'to_date': str(c.to_date)
            }
            for c in simulation.customers.all()
        ]
    }
    
    response = JsonResponse(data)
    response['Content-Disposition'] = f'attachment; filename=LNG_Simulation_{simulation.name}.json'
    return response


@login_required
def import_json(request, simulation_id):
    """Import data from JSON"""
    simulation = get_object_or_404(Simulation, pk=simulation_id, user=request.user)
    
    if request.method == 'POST':
        form = JSONUploadForm(request.POST, request.FILES)
        if form.is_valid():
            json_file = request.FILES['json_file']
            try:
                data = json.loads(json_file.read().decode('utf-8'))
                
                # Clear existing data
                simulation.suppliers.all().delete()
                simulation.cargos.all().delete()
                simulation.customers.all().delete()
                
                # Import suppliers
                if 'suppliers' in data:
                    for s in data['suppliers']:
                        Supplier.objects.create(
                            simulation=simulation,
                            name=s['name'],
                            daily_supply=s['daily_supply'],
                            from_date=s['from_date'],
                            to_date=s['to_date']
                        )
                
                # Import cargos
                if 'cargos' in data:
                    for c in data['cargos']:
                        Cargo.objects.create(
                            simulation=simulation,
                            cargo_name=c['cargo_name'],
                            delivery_date=c['date'],
                            amount=c['amount']
                        )
                
                # Import customers
                if 'customers' in data:
                    for c in data['customers']:
                        Customer.objects.create(
                            simulation=simulation,
                            name=c['name'],
                            daily_demand=c['daily_demand'],
                            from_date=c['from_date'],
                            to_date=c['to_date']
                        )
                
                messages.success(request, 'Data imported successfully!')
                return redirect('lng_planner:dashboard')
                
            except Exception as e:
                messages.error(request, f'Error importing JSON: {str(e)}')
    else:
        form = JSONUploadForm()
    
    return render(request, 'lng_planner/import_json.html', {
        'form': form,
        'simulation': simulation
    })


# HELPER FUNCTIONS

def get_simulation_data(simulation):
    """Get all data for a simulation"""
    daily_data = calculate_daily_data(simulation)
    plants = Plant.objects.filter(
        id__in=simulation.plant_inventories.values_list('plant_id', flat=True)
    )
    
    # Find first negative date for each plant
    plant_alerts = {}
    for plant in plants:
        first_negative = next((d for d in daily_data if d['plant_data'].get(plant.id, {}).get('is_negative', False)), None)
        if first_negative:
            plant_alerts[plant.id] = {
                'plant': plant,
                'date': first_negative['date'],
                'inventory': first_negative['plant_data'][plant.id]['inventory']
            }
    
    return {
        'suppliers': simulation.suppliers.all(),
        'cargos': simulation.cargos.all(),
        'customers': simulation.customers.all(),
        'plants': plants,
        'daily_data': daily_data,
        'plant_alerts': plant_alerts,
        'total_days': len(daily_data)
    }


def calculate_daily_data(simulation):
    """Calculate daily inventory projection for each plant"""
    start_date = simulation.start_date
    end_date = simulation.end_date
    
    # Initialize plant inventories
    plant_inventories = {}
    for pi in simulation.plant_inventories.all():
        plant_inventories[pi.plant_id] = float(pi.opening_inventory)
    
    daily_data = []
    current_date = start_date
    
    while current_date <= end_date:
        day_data = {
            'date': current_date,
            'plant_data': {},
            'total_supply': 0,
            'total_demand': 0,
            'total_inventory': 0
        }
        
        # Calculate for each plant
        for plant_id, inventory in plant_inventories.items():
            plant_supply = 0
            plant_demand = 0
            supplies = []
            demands = []
            
            # Suppliers for this plant
            for supplier in simulation.suppliers.filter(plant_id=plant_id):
                if supplier.from_date <= current_date <= supplier.to_date:
                    amount = float(supplier.daily_supply)
                    supplies.append({
                        'type': 'supplier',
                        'name': supplier.name,
                        'amount': amount
                    })
                    plant_supply += amount
            
            # Cargos for this plant
            for cargo in simulation.cargos.filter(plant_id=plant_id, delivery_date=current_date):
                amount = float(cargo.amount)
                supplies.append({
                    'type': 'cargo',
                    'name': cargo.cargo_name,
                    'amount': amount
                })
                plant_supply += amount
            
            # Customers for this plant
            for customer in simulation.customers.filter(plant_id=plant_id):
                if customer.from_date <= current_date <= customer.to_date:
                    amount = float(customer.daily_demand)
                    demands.append({
                        'customer': customer.name,
                        'amount': amount
                    })
                    plant_demand += amount
            
            # Update plant inventory
            new_inventory = inventory + plant_supply - plant_demand
            plant_inventories[plant_id] = new_inventory
            
            day_data['plant_data'][plant_id] = {
                'supplies': supplies,
                'demands': demands,
                'supply': plant_supply,
                'demand': plant_demand,
                'inventory': new_inventory,
                'is_negative': new_inventory < 0
            }
            
            day_data['total_supply'] += plant_supply
            day_data['total_demand'] += plant_demand
            day_data['total_inventory'] += new_inventory
        
        daily_data.append(day_data)
        current_date += timedelta(days=1)
    
    return daily_data


def initialize_sample_data(simulation):
    """Initialize simulation with sample data"""
    # Get or create sample plants if they don't exist
    plants = Plant.objects.all()
    if not plants.exists():
        return
    
    # Use first two plants for sample data
    plant_list = list(plants[:2])
    if len(plant_list) < 2:
        plant_list = plant_list * 2  # Duplicate if only one plant
    
    # Sample suppliers
    Supplier.objects.create(
        simulation=simulation,
        plant=plant_list[0],
        name='Supplier A',
        daily_supply=100,
        from_date=simulation.start_date,
        to_date=simulation.end_date
    )
    Supplier.objects.create(
        simulation=simulation,
        plant=plant_list[1],
        name='Supplier B',
        daily_supply=120,
        from_date=simulation.start_date,
        to_date=simulation.start_date + timedelta(days=180)
    )
    
    # Sample cargos
    Cargo.objects.create(
        simulation=simulation,
        plant=plant_list[0],
        cargo_name='Cargo-001',
        delivery_date=simulation.start_date + timedelta(days=15),
        amount=5000
    )
    Cargo.objects.create(
        simulation=simulation,
        plant=plant_list[1],
        cargo_name='Cargo-002',
        delivery_date=simulation.start_date + timedelta(days=80),
        amount=4500
    )
    
    # Sample customers
    Customer.objects.create(
        simulation=simulation,
        plant=plant_list[0],
        name='Customer X',
        daily_demand=80,
        from_date=simulation.start_date,
        to_date=simulation.end_date
    )
    Customer.objects.create(
        simulation=simulation,
        plant=plant_list[1],
        name='Customer Y',
        daily_demand=70,
        from_date=simulation.start_date,
        to_date=simulation.start_date + timedelta(days=180)
    )


# MASTER SIMULATION FUNCTIONS

@login_required
def manage_master_simulation(request):
    """Admin view to create/edit master simulation"""
    if not request.user.is_staff:
        messages.error(request, 'Only administrators can manage master simulation')
        return redirect('lng_planner:dashboard')
    
    master = Simulation.objects.filter(is_master=True).first()
    
    if request.method == 'POST':
        form = MasterSimulationForm(request.POST, instance=master)
        if form.is_valid():
            master_sim = form.save(commit=False)
            master_sim.user = None  # Master has no user
            master_sim.is_master = True
            master_sim.is_active = False  # Master is never active
            master_sim.save()
            
            messages.success(request, 'Master simulation saved successfully!')
            return redirect('lng_planner:setup_plants', simulation_id=master_sim.id)
    else:
        form = MasterSimulationForm(instance=master)
    
    return render(request, 'lng_planner/manage_master.html', {
        'form': form,
        'master': master,
        'title': 'Edit Master Simulation' if master else 'Create Master Simulation'
    })


@login_required
def copy_from_master(request):
    """Copy master simulation to user's own simulation"""
    master = Simulation.objects.filter(is_master=True).first()
    
    if not master:
        messages.error(request, 'No master simulation found!')
        return redirect('lng_planner:dashboard')
    
    # Create new simulation for user
    new_sim = Simulation.objects.create(
        user=request.user,
        name=f"{master.name} - Copy {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        start_date=master.start_date,
        end_date=master.end_date,
        is_master=False,
        is_active=True
    )
    
    # Deactivate other user simulations
    Simulation.objects.filter(user=request.user).exclude(pk=new_sim.pk).update(is_active=False)
    
    # Copy plant inventories
    for pi in master.plant_inventories.all():
        PlantInventory.objects.create(
            simulation=new_sim,
            plant=pi.plant,
            opening_inventory=pi.opening_inventory
        )
    
    # Copy suppliers
    for supplier in master.suppliers.all():
        Supplier.objects.create(
            simulation=new_sim,
            plant=supplier.plant,
            name=supplier.name,
            daily_supply=supplier.daily_supply,
            from_date=supplier.from_date,
            to_date=supplier.to_date
        )
    
    # Copy cargos
    for cargo in master.cargos.all():
        Cargo.objects.create(
            simulation=new_sim,
            plant=cargo.plant,
            cargo_name=cargo.cargo_name,
            delivery_date=cargo.delivery_date,
            amount=cargo.amount
        )
    
    # Copy customers
    for customer in master.customers.all():
        Customer.objects.create(
            simulation=new_sim,
            plant=customer.plant,
            name=customer.name,
            daily_demand=customer.daily_demand,
            from_date=customer.from_date,
            to_date=customer.to_date
        )
    
    messages.success(request, f'Created new simulation from master: {new_sim.name}')
    return redirect('lng_planner:dashboard')


@login_required
def refresh_master_from_sap(request):
    """Refresh master simulation from SAP API (admin only)"""
    if not request.user.is_staff:
        messages.error(request, 'Only administrators can refresh master data')
        return redirect('lng_planner:dashboard')
    
    master = Simulation.objects.filter(is_master=True).first()
    
    if not master:
        messages.error(request, 'No master simulation found. Please create one first.')
        return redirect('lng_planner:manage_master')
    
    if not master.sap_api_url:
        messages.error(request, f'SAP API URL not configured for master simulation. Please update it.')
        return redirect('lng_planner:manage_master')
    
    try:
        # Fetch data from SAP
        messages.info(request, f'Fetching data from SAP: {master.sap_api_url}')
        response = requests.get(master.sap_api_url, timeout=30)
        response.raise_for_status()
        sap_data = response.json()
        
        # Clear existing master data
        deleted_suppliers = master.suppliers.all().delete()
        deleted_cargos = master.cargos.all().delete()
        deleted_customers = master.customers.all().delete()
        deleted_inventories = master.plant_inventories.all().delete()
        
        counts = {
            'plants': 0,
            'suppliers': 0,
            'cargos': 0,
            'customers': 0,
            'errors': []
        }
        
        # Import plant inventories
        if 'plant_inventories' in sap_data:
            for pi_data in sap_data['plant_inventories']:
                try:
                    plant = Plant.objects.filter(name=pi_data['plant_name']).first()
                    if plant:
                        PlantInventory.objects.create(
                            simulation=master,
                            plant=plant,
                            opening_inventory=pi_data['opening_inventory']
                        )
                        counts['plants'] += 1
                    else:
                        counts['errors'].append(f"Plant not found: {pi_data['plant_name']}")
                except Exception as e:
                    counts['errors'].append(f"Plant inventory error: {str(e)}")
        
        # Import suppliers
        if 'suppliers' in sap_data:
            for s in sap_data['suppliers']:
                try:
                    plant = Plant.objects.filter(name=s['plant_name']).first()
                    if plant:
                        Supplier.objects.create(
                            simulation=master,
                            plant=plant,
                            name=s['name'],
                            daily_supply=s['daily_supply'],
                            from_date=s['from_date'],
                            to_date=s['to_date']
                        )
                        counts['suppliers'] += 1
                    else:
                        counts['errors'].append(f"Plant not found for supplier: {s['name']} -> {s['plant_name']}")
                except Exception as e:
                    counts['errors'].append(f"Supplier error ({s.get('name', 'unknown')}): {str(e)}")
        
        # Import cargos
        if 'cargos' in sap_data:
            for c in sap_data['cargos']:
                try:
                    plant = Plant.objects.filter(name=c['plant_name']).first()
                    if plant:
                        Cargo.objects.create(
                            simulation=master,
                            plant=plant,
                            cargo_name=c['cargo_name'],
                            delivery_date=c['date'],
                            amount=c['amount']
                        )
                        counts['cargos'] += 1
                    else:
                        counts['errors'].append(f"Plant not found for cargo: {c['cargo_name']} -> {c['plant_name']}")
                except Exception as e:
                    counts['errors'].append(f"Cargo error ({c.get('cargo_name', 'unknown')}): {str(e)}")
        
        # Import customers
        if 'customers' in sap_data:
            for c in sap_data['customers']:
                try:
                    plant = Plant.objects.filter(name=c['plant_name']).first()
                    if plant:
                        Customer.objects.create(
                            simulation=master,
                            plant=plant,
                            name=c['name'],
                            daily_demand=c['daily_demand'],
                            from_date=c['from_date'],
                            to_date=c['to_date']
                        )
                        counts['customers'] += 1
                    else:
                        counts['errors'].append(f"Plant not found for customer: {c['name']} -> {c['plant_name']}")
                except Exception as e:
                    counts['errors'].append(f"Customer error ({c.get('name', 'unknown')}): {str(e)}")
        
        master.last_sap_sync = datetime.now()
        master.save()
        
        # Build success message
        success_msg = f'✓ Master simulation refreshed from SAP! '
        success_msg += f'Imported: {counts["plants"]} plant inventories, '
        success_msg += f'{counts["suppliers"]} suppliers, '
        success_msg += f'{counts["cargos"]} cargos, '
        success_msg += f'{counts["customers"]} customers. '
        success_msg += f'Last sync: {master.last_sap_sync.strftime("%Y-%m-%d %H:%M:%S")}'
        
        messages.success(request, success_msg)
        
        # Show any errors as warnings
        if counts['errors']:
            for error in counts['errors'][:5]:  # Show first 5 errors
                messages.warning(request, f'⚠️ {error}')
            if len(counts['errors']) > 5:
                messages.warning(request, f'⚠️ ... and {len(counts["errors"]) - 5} more errors')
        
    except requests.Timeout:
        messages.error(request, f'⏱️ Timeout: SAP API took too long to respond (>30 seconds)')
    except requests.ConnectionError:
        messages.error(request, f'🔌 Connection Error: Could not reach SAP API at {master.sap_api_url}')
    except requests.HTTPError as e:
        messages.error(request, f'❌ HTTP Error: SAP API returned error {e.response.status_code}')
    except json.JSONDecodeError:
        messages.error(request, f'📄 Invalid JSON: SAP API did not return valid JSON data')
    except Exception as e:
        messages.error(request, f'❌ Unexpected error: {str(e)}')
    
    return redirect('lng_planner:dashboard')

def mock_sap_api(request):
    """Mock SAP API endpoint for testing"""
    # This returns test data in the expected format
    return JsonResponse({
        "plant_inventories": [
            {
                "plant_name": "Dahej",
                "opening_inventory": 12000
            },
            {
                "plant_name": "Ennore",
                "opening_inventory": 9000
            }
        ],
        "suppliers": [
            {
                "name": "Supplier A",
                "plant_name": "Dahej",
                "daily_supply": 150,
                "from_date": "2026-01-01",
                "to_date": "2026-12-31"
            },
            {
                "name": "Supplier B",
                "plant_name": "Ennore",
                "daily_supply": 130,
                "from_date": "2026-01-01",
                "to_date": "2026-06-30"
            }
        ],
        "cargos": [
            {
                "cargo_name": "CARGO1",
                "plant_name": "Dahej",
                "date": "2026-02-15",
                "amount": 6000
            },
            {
                "cargo_name": "CARGO2",
                "plant_name": "Ennore",
                "date": "2026-03-10",
                "amount": 5500
            }
        ],
        "customers": [
            {
                "name": "Customer 1",
                "plant_name": "Dahej",
                "daily_demand": 90,
                "from_date": "2026-01-01",
                "to_date": "2026-12-31"
            },
            {
                "name": "Customer 2",
                "plant_name": "Ennore",
                "daily_demand": 75,
                "from_date": "2026-01-01",
                "to_date": "2026-08-31"
            }
        ]
    })

