from datetime import date
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import RequestFactory, TestCase
from openpyxl import load_workbook

from .models import Plant, Simulation, PlantInventory, Supplier, Customer
from .views import export_excel


class ExportExcelBacklogTest(TestCase):
    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username='testuser', password='password')
        self.factory = RequestFactory()

        self.plant = Plant.objects.create(name='Test Plant')
        self.simulation = Simulation.objects.create(
            user=self.user,
            name='Backlog Simulation',
            start_date=date(2026, 1, 1),
            end_date=date(2026, 1, 1)
        )
        PlantInventory.objects.create(
            simulation=self.simulation,
            plant=self.plant,
            opening_inventory=10
        )
        Supplier.objects.create(
            simulation=self.simulation,
            plant=self.plant,
            name='Test Supplier',
            daily_supply=5,
            from_date=self.simulation.start_date,
            to_date=self.simulation.end_date
        )
        Customer.objects.create(
            simulation=self.simulation,
            plant=self.plant,
            name='Mahesh Prajapati',
            daily_demand=9,
            from_date=self.simulation.start_date,
            to_date=self.simulation.end_date
        )

    def test_export_excel_closing_inventory_shows_backlog(self):
        request = self.factory.get('/export_excel/')
        request.user = self.user

        response = export_excel(request, self.simulation.id)
        workbook = load_workbook(filename=BytesIO(response.content), data_only=True)
        worksheet = workbook.active

        closing_row = None
        for row in worksheet.iter_rows(values_only=True):
            if row and row[0] == f'    Closing Inventory — {self.plant.name}':
                closing_row = row
                break

        self.assertIsNotNone(closing_row, 'Closing Inventory row not found in exported workbook')
        self.assertEqual(closing_row[1], '0 (4)')
