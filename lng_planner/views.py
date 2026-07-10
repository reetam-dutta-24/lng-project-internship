from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from datetime import datetime, timedelta
from decimal import Decimal
import json
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Simulation, SimulationComment, Supplier, SupplierDate, Cargo, Customer, CustomerDate, Refinery, RefineryDate, Plant, PlantInventory, APIConfiguration, MasterVersion
from .forms import (
    SimulationForm, SupplierForm, SupplierDateForm, CargoForm, CustomerForm, CustomerDateForm, 
    RefineryForm, RefineryDateForm, PlantForm, PlantInventoryForm,
    APIConfigurationForm, JSONUploadForm, MasterSimulationForm
)


def landing_page(request):
    """Landing page view - public homepage"""
    return render(request, 'lng_planner/landing.html')


# Helper function to check if user is Planning Admin
def is_planning_admin(user):
    """Check if user has Planning Admin role (via group or superuser)"""
    if not user.is_authenticated:
        return False
    # Superusers have all permissions
    if user.is_superuser:
        return True
    # Check if user is in Planning Admin group
    return user.groups.filter(name='Planning Admin').exists()


# Helper function to get next version number
def get_next_version_number():
    """Generate the next master version number"""
    from django.db.models import Max
    
    # Get all versions and find the maximum numeric value
    versions = MasterVersion.objects.all()
    max_num = 0
    
    for v in versions:
        try:
            num = int(v.version_number[1:])  # Extract number after 'V'
            if num > max_num:
                max_num = num
        except (ValueError, IndexError):
            continue
    
    next_num = max_num + 1
    return f"V{next_num}"


def copy_simulation_data(source_simulation, target_simulation):
    """
    Copy all data from source simulation to target simulation.
    Used for creating master simulations from user simulations or SAP data.
    
    Returns a dictionary with counts of copied items for logging.
    """
    # Clear existing data in target (except plant inventories which we'll replace)
    SupplierDate.objects.filter(supplier__simulation=target_simulation).delete()
    CustomerDate.objects.filter(customer__simulation=target_simulation).delete()
    RefineryDate.objects.filter(refinery__simulation=target_simulation).delete()
    Supplier.objects.filter(simulation=target_simulation).delete()
    Customer.objects.filter(simulation=target_simulation).delete()
    Refinery.objects.filter(simulation=target_simulation).delete()
    Cargo.objects.filter(simulation=target_simulation).delete()
    PlantInventory.objects.filter(simulation=target_simulation).delete()
    SimulationComment.objects.filter(simulation=target_simulation).delete()
    
    # Copy Suppliers with their date ranges
    suppliers_copied = 0
    for supplier in source_simulation.suppliers.all():
        new_supplier = Supplier.objects.create(
            simulation=target_simulation,
            plant=supplier.plant,
            name=supplier.name,
            preference=supplier.preference
        )
        
        # Copy date ranges
        date_ranges_copied = 0
        for date_range in supplier.date_ranges.all():
            SupplierDate.objects.create(
                supplier=new_supplier,
                from_date=date_range.from_date,
                to_date=date_range.to_date,
                daily_supply=date_range.daily_supply
            )
            date_ranges_copied += 1
        
        suppliers_copied += 1
    
    # Copy Customers with their date ranges
    customers_copied = 0
    for customer in source_simulation.customers.all():
        new_customer = Customer.objects.create(
            simulation=target_simulation,
            plant=customer.plant,
            name=customer.name,
            preference=customer.preference
        )
        
        # Copy date ranges with supplier allocation
        customer_dates_copied = 0
        for date_range in customer.date_ranges.all():
            CustomerDate.objects.create(
                customer=new_customer,
                supplier=date_range.supplier,
                from_date=date_range.from_date,
                to_date=date_range.to_date,
                daily_demand=date_range.daily_demand
            )
            customer_dates_copied += 1
        
        customers_copied += 1
    
    # Copy Refineries with their date ranges
    refineries_copied = 0
    for refinery in source_simulation.refineries.all():
        new_refinery = Refinery.objects.create(
            simulation=target_simulation,
            plant=refinery.plant,
            name=refinery.name,
            preference=refinery.preference
        )
        
        # Copy date ranges
        refinery_dates_copied = 0
        for date_range in refinery.date_ranges.all():
            RefineryDate.objects.create(
                refinery=new_refinery,
                from_date=date_range.from_date,
                to_date=date_range.to_date,
                daily_refinery_demand=date_range.daily_refinery_demand
            )
            refinery_dates_copied += 1
        
        refineries_copied += 1
    
    # Copy Cargos
    cargos_copied = 0
    for cargo in source_simulation.cargos.all():
        Cargo.objects.create(
            simulation=target_simulation,
            plant=cargo.plant,
            cargo_name=cargo.cargo_name,
            delivery_date=cargo.delivery_date,
            amount=cargo.amount
        )
        cargos_copied += 1
    
    # Copy Plant Inventories
    inventories_copied = 0
    for pi in source_simulation.plant_inventories.all():
        PlantInventory.objects.create(
            simulation=target_simulation,
            plant=pi.plant,
            opening_inventory=pi.opening_inventory
        )
        inventories_copied += 1
    
    # Copy Simulation Comments
    comments_copied = 0
    for comment in source_simulation.comments.all():
        SimulationComment.objects.create(
            simulation=target_simulation,
            comment=comment.comment,
            created_by=comment.created_by
        )
        comments_copied += 1
    
    return {
        'suppliers': suppliers_copied,
        'supplier_dates': sum(s.date_ranges.count() for s in source_simulation.suppliers.all()),
        'customers': customers_copied,
        'customer_dates': sum(c.date_ranges.count() for c in source_simulation.customers.all()),
        'refineries': refineries_copied,
        'refinery_dates': sum(r.date_ranges.count() for r in source_simulation.refineries.all()),
        'cargos': cargos_copied,
        'inventories': inventories_copied,
        'comments': comments_copied,
    }


# ── Master Version Management Views ────────────────────────────────────────
@login_required
@user_passes_test(is_planning_admin)
def refresh_master_from_sap(request):
    """Refresh master data from SAP and create a new master version"""
    current_master = Simulation.objects.filter(is_master=True).first()
    
    if not current_master:
        messages.error(request, 'No master simulation found. Please contact administrator.')
        return redirect('lng_planner:dashboard')
    
    # Get SAP API configuration
    sap_api_url = current_master.sap_api_url or APIConfiguration.objects.filter(key='sap_api_url').first()
    if sap_api_url:
        sap_api_url = getattr(sap_api_url, 'value', None) or sap_api_url
    
    if not sap_api_url:
        messages.error(request, 'SAP API URL not configured. Please set it in the master simulation settings.')
        return redirect('lng_planner:dashboard')
    
    try:
        # === LOGGING: Old Active Master Version ===
        old_master_version = None
        if hasattr(current_master, 'master_version') and current_master.master_version:
            old_master_version = current_master.master_version
        
        # Fetch data from SAP (mock implementation - replace with actual API call)
        response = requests.get(sap_api_url, timeout=30)
        sap_data = response.json()
        
        # Step 1: Create new master version
        version_number = get_next_version_number()
        new_master_version = MasterVersion.objects.create(
            version_number=version_number,
            name=f"SAP Sync - {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            source_type='SAP',
            created_by=request.user,
            description=f"Data synced from SAP API",
            is_active=True
        )
        
        # === LOGGING: New Master Version Created ===
        print(f"\n{'='*60}")
        print("SAP REFRESH - AUDIT LOG")
        print(f"{'='*60}")
        if old_master_version:
            print(f"Old Active Master Version: {old_master_version.version_number} ({old_master_version.name})")
        else:
            print("Old Active Master Version: None (initial refresh)")
        print(f"New Master Version Created: {new_master_version.version_number} ({new_master_version.name})")
        print(f"Source: SAP API ({sap_api_url})")
        
        # Step 2: Deactivate previous master versions
        deactivated_count = MasterVersion.objects.filter(is_active=True).exclude(pk=new_master_version.pk).update(is_active=False)
        print(f"Deactivated {deactivated_count} previous master version(s)")
        
        # Step 3: Deactivate old master simulation
        if current_master:
            current_master.is_master = False
            current_master.save()
            print(f"Deactivated old master simulation: {current_master.name} (ID: {current_master.id})")
        
        # Step 4: Create NEW master simulation with SAP data
        from django.utils import timezone
        new_master_simulation = Simulation.objects.create(
            name=f"Master - {version_number}",
            start_date=current_master.start_date,
            end_date=current_master.end_date,
            is_master=True,
            user=None,  # Master simulations have no owner
            master_version=new_master_version,
            sap_api_url=sap_api_url,
            last_sap_sync=timezone.now()
        )
        print(f"New Master Simulation Created: {new_master_simulation.name} (ID: {new_master_simulation.id})")
        print(f"  - is_master=True")
        print(f"  - user=NULL")
        print(f"  - master_version={new_master_version.version_number}")
        
        # Step 5: Process SAP data and populate the new master simulation
        # TODO: Implement actual SAP data processing logic here
        # For now, we'll copy existing data as a placeholder to maintain functionality
        # In production, this would parse sap_data and create Suppliers, Customers, Cargos
        
        print(f"\n⚠️  SAP Data Processing: Using placeholder (copy from old master)")
        print(f"   TODO: Implement actual SAP API parsing")
        
        # Copy data from the deactivated master as placeholder
        copy_counts = copy_simulation_data(current_master, new_master_simulation)
        
        print(f"\nData in New Master (placeholder):")
        print(f"  - Plant Inventories: {copy_counts['inventories']}")
        print(f"  - Suppliers: {copy_counts['suppliers']}")
        print(f"    └─ Supplier Date Ranges: {copy_counts['supplier_dates']}")
        print(f"  - Customers: {copy_counts['customers']}")
        print(f"    └─ Customer Date Ranges: {copy_counts['customer_dates']}")
        print(f"  - Refineries: {copy_counts['refineries']}")
        print(f"    └─ Refinery Date Ranges: {copy_counts['refinery_dates']}")
        print(f"  - Cargos: {copy_counts['cargos']}")
        
        # === LOGGING: New Active Master Version ===
        new_active = Simulation.objects.filter(is_master=True).first()
        new_active_version = new_active.master_version if new_active and hasattr(new_active, 'master_version') else None
        print(f"\nNew Active Master Version: {new_active_version.version_number if new_active_version else 'None'}")
        print(f"{'='*60}\n")
        
        messages.success(request, f'Master version {version_number} created successfully from SAP! All users will now use this as the active master.')
        
    except Exception as e:
        print(f"\n❌ ERROR during SAP refresh: {str(e)}")
        print(f"{'='*60}\n")
        messages.error(request, f'Error refreshing from SAP: {str(e)}')
    
    return redirect('lng_planner:dashboard')


@login_required
@user_passes_test(is_planning_admin)
def publish_simulation_to_master(request, simulation_id):
    """Publish simulation data as a new master version"""
    simulation = get_object_or_404(Simulation, pk=simulation_id, user=request.user, is_master=False)
    
    if request.method == 'POST':
        # === LOGGING: Old Active Master Version ===
        old_master_simulation = Simulation.objects.filter(is_master=True).first()
        old_master_version = None
        if old_master_simulation and hasattr(old_master_simulation, 'master_version') and old_master_simulation.master_version:
            old_master_version = old_master_simulation.master_version
        
        # Step 1: Create new master version
        version_number = get_next_version_number()
        new_master_version = MasterVersion.objects.create(
            version_number=version_number,
            name=f"Published from {simulation.name}",
            source_type='SIMULATION',
            created_by=request.user,
            description=f"Published from simulation: {simulation.name} by {request.user.get_full_name() or request.user.username}",
            is_active=True,
            source_simulation=simulation
        )
        
        # === LOGGING: New Master Version Created ===
        print(f"\n{'='*60}")
        print("PUBLISH TO MASTER - AUDIT LOG")
        print(f"{'='*60}")
        if old_master_version:
            print(f"Old Active Master Version: {old_master_version.version_number} ({old_master_version.name})")
        else:
            print("Old Active Master Version: None (initial publish)")
        print(f"New Master Version Created: {new_master_version.version_number} ({new_master_version.name})")
        print(f"Source Simulation: {simulation.name} (ID: {simulation.id}, User: {simulation.user.username})")
        
        # Step 2: Deactivate previous master versions
        deactivated_count = MasterVersion.objects.filter(is_active=True).exclude(pk=new_master_version.pk).update(is_active=False)
        print(f"Deactivated {deactivated_count} previous master version(s)")
        
        # Step 3: Deactivate old master simulation (if exists)
        if old_master_simulation:
            old_master_simulation.is_master = False
            old_master_simulation.save()
            print(f"Deactivated old master simulation: {old_master_simulation.name} (ID: {old_master_simulation.id})")
        
        # Step 4: Create NEW master simulation with the published data
        from django.utils import timezone
        new_master_simulation = Simulation.objects.create(
            name=f"Master - {version_number}",
            start_date=simulation.start_date,
            end_date=simulation.end_date,
            is_master=True,
            user=None,  # Master simulations have no owner
            master_version=new_master_version,
            sap_api_url=simulation.sap_api_url,
            last_sap_sync=timezone.now()
        )
        print(f"New Master Simulation Created: {new_master_simulation.name} (ID: {new_master_simulation.id})")
        print(f"  - is_master=True")
        print(f"  - user=NULL")
        print(f"  - master_version={new_master_version.version_number}")
        
        # Step 5: Copy all data from published simulation to new master simulation
        try:
            copy_counts = copy_simulation_data(simulation, new_master_simulation)
            
            # === LOGGING: Data Copied ===
            print(f"\nData Copied to New Master:")
            print(f"  - Plant Inventories: {copy_counts['inventories']}")
            print(f"  - Suppliers: {copy_counts['suppliers']}")
            print(f"    └─ Supplier Date Ranges: {copy_counts['supplier_dates']}")
            print(f"  - Customers: {copy_counts['customers']}")
            print(f"    └─ Customer Date Ranges: {copy_counts['customer_dates']}")
            print(f"  - Refineries: {copy_counts['refineries']}")
            print(f"    └─ Refinery Date Ranges: {copy_counts['refinery_dates']}")
            print(f"  - Cargos: {copy_counts['cargos']}")
            print(f"  - Comments: {copy_counts['comments']}")
            
            # === LOGGING: New Active Master Version ===
            new_active = Simulation.objects.filter(is_master=True).first()
            new_active_version = new_active.master_version if new_active and hasattr(new_active, 'master_version') else None
            print(f"\nNew Active Master Version: {new_active_version.version_number if new_active_version else 'None'}")
            print(f"{'='*60}\n")
            
            messages.success(request, f'Master version {version_number} created successfully! All data copied. Old master was {old_master_version.version_number if old_master_version else "none"}.')
        except Exception as e:
            # Rollback if copying fails
            new_master_simulation.delete()
            new_master_version.is_active = False
            new_master_version.save()
            print(f"\n❌ ERROR during publish: {str(e)}")
            print(f"{'='*60}\n")
            messages.error(request, f'Error copying data to master: {str(e)}')
        
        return redirect('lng_planner:dashboard')
    
    # Show confirmation page
    return render(request, 'lng_planner/publish_confirm.html', {
        'simulation': simulation,
        'next_version': get_next_version_number()
    })


@login_required
def view_master_history(request):
    """View master version history"""
    master_versions = MasterVersion.objects.all()
    
    return render(request, 'lng_planner/master_history.html', {
        'master_versions': master_versions
    })


# ── Dashboard and simulation selection ─────────────────────────────────────
@login_required
def dashboard(request):
    # Main dashboard view for the current user. Chooses an active simulation,
    # falls back to the master simulation if no active one is available,
    # and populates the context for rendering the dashboard page.
    from datetime import datetime
    
    master_simulation = Simulation.objects.filter(is_master=True).first()
    
    # Determine simulation visibility based on role
    is_super_user = request.user.is_superuser
    is_planning_admin_role = is_super_user or request.user.groups.filter(name='Planning Admin').exists()
    
    # ONLY superuser can see all simulations
    # Both Planner and Planning Admin only see their own simulations
    if is_super_user:
        # Superuser sees ALL non-master simulations
        simulations = Simulation.objects.filter(is_master=False).order_by('-created_at')
    else:
        # Planners and Planning Admins can only see their own simulations
        simulations = Simulation.objects.filter(user=request.user, is_master=False).order_by('-created_at')
    
    # Get the active simulation for this user (must be non-master)
    active_simulation = simulations.filter(is_active=True).first()
    
    # If no active simulation found, default to master (read-only mode)
    if not active_simulation:
        if master_simulation:
            active_simulation = master_simulation
        elif simulations.exists():
            # Fallback: activate the first available simulation
            active_simulation = simulations.first()
            active_simulation.is_active = True
            active_simulation.save()
    
    # Get current active master version
    current_master_version = None
    if master_simulation and hasattr(master_simulation, 'master_version') and master_simulation.master_version:
        current_master_version = master_simulation.master_version
    
    # Handle date range filter from GET parameters
    filter_from_date = None
    filter_to_date = None
    filter_error = None
    
    if request.GET.get('from_date') and request.GET.get('to_date'):
        try:
            filter_from_date = datetime.strptime(request.GET['from_date'], '%Y-%m-%d').date()
            filter_to_date = datetime.strptime(request.GET['to_date'], '%Y-%m-%d').date()
            
            # Validate date range
            if filter_from_date > filter_to_date:
                filter_error = "From Date cannot be after To Date"
            elif filter_from_date < active_simulation.start_date or filter_to_date > active_simulation.end_date:
                filter_error = f"Selected range must be within {active_simulation.start_date} to {active_simulation.end_date}"
        except ValueError:
            filter_error = "Invalid date format"
    
    context = {
        'simulations': simulations,
        'active_simulation': active_simulation,
        'master_simulation': master_simulation,
        'is_viewing_master': active_simulation and active_simulation.is_master if active_simulation else False,
        'current_master_version': current_master_version,
        'is_planning_admin': is_planning_admin_role,
        'is_super_user': is_super_user,
        'filter_from_date': filter_from_date,
        'filter_to_date': filter_to_date,
        'filter_error': filter_error,
    }
    
    if active_simulation:
        # Pass filter parameters to get_simulation_data
        context.update(get_simulation_data(active_simulation, filter_from_date, filter_to_date))
    
    return render(request, 'lng_planner/dashboard.html', context)


@login_required
def bulk_supply_adjustment(request):
    """Handle bulk supply adjustment for a supplier and its linked customers/refineries"""
    from django.db import transaction
    
    if request.method == 'POST':
        simulation_id = request.POST.get('simulation_id')
        supplier_id = request.POST.get('supplier_id')
        adjustment_type = request.POST.get('adjustment_type')  # 'increase' or 'decrease'
        percentage_str = request.POST.get('percentage', '0')
        
        try:
            percentage = float(percentage_str)
        except (ValueError, TypeError):
            return JsonResponse({'success': False, 'error': 'Invalid percentage value'})
        
        # Validate inputs
        if not simulation_id or not supplier_id:
            return JsonResponse({'success': False, 'error': 'Missing required parameters'})
        
        if adjustment_type not in ['increase', 'decrease']:
            return JsonResponse({'success': False, 'error': 'Invalid adjustment type'})
        
        if percentage < 0 or percentage > 100:
            return JsonResponse({'success': False, 'error': 'Percentage must be between 0 and 100'})
        
        try:
            simulation = Simulation.objects.get(pk=simulation_id)
            supplier = Supplier.objects.get(pk=supplier_id, simulation=simulation)
        except (Simulation.DoesNotExist, Supplier.DoesNotExist):
            return JsonResponse({'success': False, 'error': 'Invalid simulation or supplier'})
        
        # Calculate the multiplier
        if adjustment_type == 'increase':
            multiplier = 1 + (percentage / 100)
        else:  # decrease
            multiplier = 1 - (percentage / 100)
        
        # Perform the update in a transaction
        try:
            with transaction.atomic():
                updated_records = {
                    'supplier_dates': 0,
                    'customer_dates': 0,
                    'refinery_dates': 0,
                }
                
                # Update SupplierDate records
                for sd in supplier.date_ranges.all():
                    old_supply = sd.daily_supply
                    sd.daily_supply = round(old_supply * multiplier, 2)
                    sd.save()
                    updated_records['supplier_dates'] += 1
                
                # Update CustomerDate records linked to this supplier
                for cd in supplier.customer_date_ranges.all():
                    old_demand = cd.daily_demand
                    cd.daily_demand = round(old_demand * multiplier, 2)
                    cd.save()
                    updated_records['customer_dates'] += 1
                
                # Update RefineryDate records linked to this supplier
                for rd in supplier.refinery_date_ranges.all():
                    old_demand = rd.daily_refinery_demand
                    rd.daily_refinery_demand = round(old_demand * multiplier, 2)
                    rd.save()
                    updated_records['refinery_dates'] += 1
                
                total_updated = sum(updated_records.values())
                
                # Log the adjustment (optional audit trail)
                from django.utils import timezone
                print(f"BULK ADJUSTMENT - User: {request.user.username}, "
                      f"Time: {timezone.now()}, Supplier: {supplier.name}, "
                      f"Type: {adjustment_type}, Percentage: {percentage}%, "
                      f"Records Updated: {total_updated}")
                
                return JsonResponse({
                    'success': True,
                    'message': f'Successfully updated {total_updated} records',
                    'details': {
                        'supplier_name': supplier.name,
                        'adjustment_type': adjustment_type,
                        'percentage': percentage,
                        'multiplier': multiplier,
                        'records_updated': updated_records,
                        'total_updated': total_updated,
                    }
                })
                
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Failed to update records: {str(e)}'
            })
    
    # GET request - show preview
    simulation_id = request.GET.get('simulation_id')
    supplier_id = request.GET.get('supplier_id')
    adjustment_type = request.GET.get('adjustment_type', 'decrease')
    percentage_str = request.GET.get('percentage', '0')
    
    try:
        percentage = float(percentage_str) if percentage_str else 0
    except (ValueError, TypeError):
        percentage = 0
    
    if not simulation_id or not supplier_id:
        return JsonResponse({'success': False, 'error': 'Missing required parameters'})
    
    try:
        simulation = Simulation.objects.get(pk=simulation_id)
        supplier = Supplier.objects.get(pk=supplier_id, simulation=simulation)
    except (Simulation.DoesNotExist, Supplier.DoesNotExist):
        return JsonResponse({'success': False, 'error': 'Invalid simulation or supplier'})
    
    # Calculate the multiplier for preview
    if adjustment_type == 'increase':
        multiplier = 1 + (percentage / 100)
    else:  # decrease
        multiplier = 1 - (percentage / 100)
    
    # Get sample data for preview
    supplier_dates = list(supplier.date_ranges.all()[:3].values('from_date', 'to_date', 'daily_supply'))
    customer_count = supplier.customer_date_ranges.count()
    refinery_count = supplier.refinery_date_ranges.count()
    
    # DEBUG: Log the counts being returned
    print(f"\n🔍 BULK ADJUSTMENT PREVIEW:")
    print(f"  Supplier: {supplier.name} (ID: {supplier.id})")
    print(f"  Simulation: {simulation.name} (ID: {simulation.id})")
    print(f"  CustomerDate count: {customer_count}")
    print(f"  RefineryDate count: {refinery_count}")
    if customer_count > 0:
        print(f"  Linked Customers:")
        for cd in supplier.customer_date_ranges.all():
            print(f"    - {cd.customer.name}: {cd.daily_demand} MT/day")
    print("=" * 60 + "\n")
    
    # Calculate old and new supply for first date range as example
    first_date_range = supplier.date_ranges.first()
    if first_date_range:
        old_supply = first_date_range.daily_supply
        new_supply = round(old_supply * multiplier, 2)
    else:
        old_supply = 0
        new_supply = 0
    
    return JsonResponse({
        'success': True,
        'preview': {
            'supplier_name': supplier.name,
            'plant_name': supplier.plant.name if supplier.plant else 'N/A',
            'adjustment_type': adjustment_type,
            'percentage': percentage,
            'multiplier': multiplier,
            'old_supply_example': old_supply,
            'new_supply_example': new_supply,
            'supplier_date_ranges_count': supplier.date_ranges.count(),
            'customer_count': customer_count,
            'refinery_count': refinery_count,
            'total_records_affected': supplier.date_ranges.count() + customer_count + refinery_count,
        }
    })


@login_required
def create_simulation(request):
    # Instantiate a new user simulation by cloning data from the master simulation.
    # Copies plant inventories, suppliers, cargos, and customers from the master.
    master = Simulation.objects.filter(is_master=True).first()
    
    if not master:
        messages.error(request, 'No master simulation found. Please contact administrator.')
        return redirect('lng_planner:dashboard')
    
    # Get current active master version
    current_master_version = None
    if hasattr(master, 'master_version') and master.master_version:
        current_master_version = master.master_version
    
    if request.method == 'POST':
        sim_name = request.POST.get('name', f"My Simulation - {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        
        new_sim = Simulation.objects.create(
            user=request.user,
            name=sim_name,
            start_date=master.start_date,
            end_date=master.end_date,
            is_master=False,
            is_active=True,
            master_version=current_master_version  # Link to current master version
        )
        
        Simulation.objects.filter(user=request.user).exclude(pk=new_sim.pk).update(is_active=False)
        
        for pi in master.plant_inventories.all():
            PlantInventory.objects.create(
                simulation=new_sim,
                plant=pi.plant,
                opening_inventory=pi.opening_inventory
            )
        
        for supplier in master.suppliers.all():
            new_supplier = Supplier.objects.create(
                simulation=new_sim,
                plant=supplier.plant,
                name=supplier.name,
                preference=supplier.preference
            )
            # Copy date ranges if they exist
            for sd in supplier.date_ranges.all():
                SupplierDate.objects.create(
                    supplier=new_supplier,
                    from_date=sd.from_date,
                    to_date=sd.to_date,
                    daily_supply=sd.daily_supply
                )
        
        for cargo in master.cargos.all():
            Cargo.objects.create(
                simulation=new_sim,
                plant=cargo.plant,
                cargo_name=cargo.cargo_name,
                delivery_date=cargo.delivery_date,
                amount=cargo.amount
            )
        
        for customer in master.customers.all():
            new_customer = Customer.objects.create(
                simulation=new_sim,
                plant=customer.plant,
                name=customer.name,
                preference=customer.preference
            )
            # Copy date ranges if they exist
            for cd in customer.date_ranges.all():
                CustomerDate.objects.create(
                    customer=new_customer,
                    from_date=cd.from_date,
                    to_date=cd.to_date,
                    daily_demand=cd.daily_demand
                )
        
        # Copy refineries with date ranges
        for refinery in master.refineries.all():
            new_refinery = Refinery.objects.create(
                simulation=new_sim,
                plant=refinery.plant,
                name=refinery.name,
                preference=refinery.preference
            )
            # Copy date ranges if they exist
            for rd in refinery.date_ranges.all():
                RefineryDate.objects.create(
                    refinery=new_refinery,
                    from_date=rd.from_date,
                    to_date=rd.to_date,
                    daily_refinery_demand=rd.daily_refinery_demand
                )
        
        messages.success(request, f'Simulation "{new_sim.name}" created from master! You can now edit it.')
        return redirect('lng_planner:dashboard')
    
    context = {
        'master': master,
        'suggested_name': f"My Simulation - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    }
    return render(request, 'lng_planner/create_from_master.html', context)


@login_required
def switch_simulation(request, simulation_id):
    # Mark the requested simulation as active for the current user.
    # Deactivates any other simulations belonging to this user.
    simulation = get_object_or_404(Simulation, pk=simulation_id)
    
    if not simulation.is_master and simulation.user != request.user:
        messages.error(request, 'You cannot access this simulation')
        return redirect('lng_planner:dashboard')
    
    Simulation.objects.filter(user=request.user).update(is_active=False)
    
    if not simulation.is_master:
        simulation.is_active = True
        simulation.save()
    
    messages.success(request, f'Switched to simulation: {simulation.name}')
    return redirect('lng_planner:dashboard')


@login_required
def delete_simulation(request, simulation_id):
    # Remove a simulation owned by the current user.
    simulation = get_object_or_404(Simulation, pk=simulation_id, user=request.user)
    simulation.delete()
    messages.success(request, 'Simulation deleted successfully!')
    return redirect('lng_planner:dashboard')


# ── Plant management views ───────────────────────────────────────────────
@login_required
def manage_plants(request):
    # Display current plants and process new plant creation requests.
    plants = Plant.objects.all()
    
    if request.method == 'POST':
        form = PlantForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Plant added successfully!')
            return redirect('lng_planner:manage_plants')
    else:
        form = PlantForm()
    
    return render(request, 'lng_planner/manage_plants.html', {
        'plants': plants,
        'form': form
    })


@login_required
def delete_plant(request, plant_id):
    # Delete plant metadata from the database.
    plant = get_object_or_404(Plant, pk=plant_id)
    plant.delete()
    messages.success(request, 'Plant deleted successfully!')
    return redirect('lng_planner:manage_plants')


@login_required
def setup_plants(request, simulation_id):
    # Configure opening inventory values for each plant in the selected simulation.
    # If plant records exist, this initializes inventory and triggers sample data.
    simulation = get_object_or_404(Simulation, pk=simulation_id, user=request.user)
    plants = Plant.objects.all()
    
    if not plants.exists():
        messages.warning(request, 'Please create plants first.')
        return redirect('lng_planner:manage_plants')
    
    if request.method == 'POST':
        for plant in plants:
            inventory = request.POST.get(f'plant_{plant.id}')
            if inventory:
                PlantInventory.objects.update_or_create(
                    simulation=simulation,
                    plant=plant,
                    defaults={'opening_inventory': float(inventory)}
                )
        
        # Populate sample suppliers, cargos, and customers after plant inventory setup.
        initialize_sample_data(simulation)
        
        messages.success(request, 'Plant inventories configured successfully!')
        return redirect('lng_planner:dashboard')
    
    existing_inventories = {
        pi.plant_id: pi.opening_inventory 
        for pi in simulation.plant_inventories.all()
    }
    
    return render(request, 'lng_planner/setup_plants.html', {
        'simulation': simulation,
        'plants': plants,
        'existing_inventories': existing_inventories
    })


# ── Supplier CRUD views ─────────────────────────────────────────────────
@login_required
def add_supplier(request, simulation_id):
    # Create a supplier record tied to the specified simulation.
    simulation = get_object_or_404(Simulation, pk=simulation_id, user=request.user)
    
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            supplier = form.save(commit=False)
            supplier.simulation = simulation
            supplier.save()
            
            # DEBUG: Print all POST keys
            print("=" * 60)
            print("DEBUG - add_supplier POST data:")
            for key, value in request.POST.items():
                if 'date_range' in key or 'daily' in key:
                    print(f"  {key} = {value}")
            print("=" * 60)
            
            # Handle additional date ranges (dynamic fields)
            date_ranges_saved = 0
            for key in request.POST:
                if key.startswith('date_ranges_') and key.endswith('_daily_supply'):
                    parts = key.split('_')
                    if len(parts) >= 4:
                        index = parts[2]  # date_ranges_1_daily_supply -> parts[2] = '1'
                        
                        daily_sup = request.POST.get(f'date_ranges_{index}_daily_supply')
                        fr_date = request.POST.get(f'date_ranges_{index}_from_date')
                        to_dt = request.POST.get(f'date_ranges_{index}_to_date')
                        
                        print(f"Processing range {index}: supply={daily_sup}, from={fr_date}, to={to_dt}")
                        
                        if daily_sup and fr_date and to_dt:
                            SupplierDate.objects.create(
                                supplier=supplier,
                                from_date=fr_date,
                                to_date=to_dt,
                                daily_supply=daily_sup
                            )
                            date_ranges_saved += 1
                            print(f"  ✓ Saved date range {index}")
            
            print(f"\nTotal date ranges saved: {date_ranges_saved}")
            print("=" * 60)
            
            messages.success(request, f'Supplier added successfully! {date_ranges_saved} date range(s) saved.')
            return redirect('lng_planner:dashboard')
    else:
        form = SupplierForm()
    
    return render(request, 'lng_planner/supplier_form.html', {
        'form': form,
        'simulation': simulation,
        'title': 'Add Supplier'
    })


@login_required
def edit_supplier(request, supplier_id):
    # Update an existing supplier entry for the current user's simulation.
    supplier = get_object_or_404(Supplier, pk=supplier_id, simulation__user=request.user)
    simulation = supplier.simulation
    
    if request.method == 'POST':
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            
            # Handle primary date range updates
            from_date = request.POST.get('from_date')
            to_date = request.POST.get('to_date')
            daily_supply = request.POST.get('daily_supply')
            date_range_id = request.POST.get('date_range_id')
            
            if from_date and to_date and daily_supply:
                if date_range_id:
                    # Update existing date range
                    date_range = get_object_or_404(SupplierDate, pk=date_range_id, supplier=supplier)
                    date_range.from_date = from_date
                    date_range.to_date = to_date
                    date_range.daily_supply = daily_supply
                    date_range.save()
                else:
                    # Create new date range
                    SupplierDate.objects.create(
                        supplier=supplier,
                        from_date=from_date,
                        to_date=to_date,
                        daily_supply=daily_supply
                    )
            
            # Handle additional date ranges (dynamic fields)
            for key in request.POST:
                if key.startswith('date_ranges_') and key.endswith('_daily_supply'):
                    parts = key.split('_')
                    if len(parts) >= 4:
                        index = parts[2]  # date_ranges_1_daily_supply -> parts[2] = '1'
                        
                        daily_sup = request.POST.get(f'date_ranges_{index}_daily_supply')
                        fr_date = request.POST.get(f'date_ranges_{index}_from_date')
                        to_dt = request.POST.get(f'date_ranges_{index}_to_date')
                        
                        if daily_sup and fr_date and to_dt:
                            # Check if this is an update (has date_range_id in POST)
                            range_id = request.POST.get(f'date_ranges_{index}_date_range_id')
                            
                            if range_id:
                                # Update existing date range
                                date_range = get_object_or_404(SupplierDate, pk=range_id, supplier=supplier)
                                date_range.from_date = fr_date
                                date_range.to_date = to_dt
                                date_range.daily_supply = daily_sup
                                date_range.save()
                            else:
                                # Create new date range
                                SupplierDate.objects.create(
                                    supplier=supplier,
                                    from_date=fr_date,
                                    to_date=to_dt,
                                    daily_supply=daily_sup
                                )
            
            messages.success(request, 'Supplier updated successfully!')
            return redirect('lng_planner:dashboard')
    else:
        form = SupplierForm(instance=supplier)
    
    # Get existing date ranges for the supplier
    date_ranges = supplier.date_ranges.all().order_by('from_date')
    
    return render(request, 'lng_planner/supplier_form.html', {
        'form': form,
        'simulation': simulation,
        'title': 'Edit Supplier',
        'supplier': supplier,
        'date_ranges': date_ranges
    })


@login_required
def delete_supplier(request, supplier_id):
    # Remove a supplier from the simulation.
    supplier = get_object_or_404(Supplier, pk=supplier_id, simulation__user=request.user)
    supplier.delete()
    messages.success(request, 'Supplier deleted successfully!')
    return redirect('lng_planner:dashboard')


@login_required
def add_supplier_date_range(request, supplier_id):
    # Add a new date range to an existing supplier
    supplier = get_object_or_404(Supplier, pk=supplier_id, simulation__user=request.user)
    
    if request.method == 'POST':
        form = SupplierDateForm(request.POST)
        if form.is_valid():
            date_range = form.save(commit=False)
            date_range.supplier = supplier
            date_range.save()
            messages.success(request, 'Date range added successfully!')
        else:
            messages.error(request, 'Error adding date range. Please check the form.')
    
    return redirect('lng_planner:dashboard')


@login_required
def edit_supplier_date_range(request, date_range_id):
    # Edit an existing supplier date range
    date_range = get_object_or_404(SupplierDate, pk=date_range_id, supplier__simulation__user=request.user)
    
    if request.method == 'POST':
        form = SupplierDateForm(request.POST, instance=date_range)
        if form.is_valid():
            form.save()
            messages.success(request, 'Date range updated successfully!')
        else:
            messages.error(request, 'Error updating date range.')
    
    return redirect('lng_planner:dashboard')


@login_required
def delete_supplier_date_range(request, date_range_id):
    # Delete a supplier date range
    date_range = get_object_or_404(SupplierDate, pk=date_range_id, supplier__simulation__user=request.user)
    supplier_name = date_range.supplier.name
    date_range.delete()
    messages.success(request, f'Date range for {supplier_name} deleted successfully!')
    return redirect('lng_planner:dashboard')


# ── Cargo CRUD views ──────────────────────────────────────────────────────
@login_required
def add_cargo(request, simulation_id):
    # Add a new cargo delivery record to the given simulation.
    simulation = get_object_or_404(Simulation, pk=simulation_id, user=request.user)
    
    if request.method == 'POST':
        form = CargoForm(request.POST)
        if form.is_valid():
            cargo = form.save(commit=False)
            cargo.simulation = simulation
            cargo.save()
            messages.success(request, 'Cargo added successfully!')
            return redirect('lng_planner:dashboard')
    else:
        form = CargoForm()
    
    return render(request, 'lng_planner/cargo_form.html', {
        'form': form,
        'simulation': simulation,
        'title': 'Add Cargo'
    })


@login_required
def edit_cargo(request, cargo_id):
    # Edit an existing cargo delivery record.
    cargo = get_object_or_404(Cargo, pk=cargo_id, simulation__user=request.user)
    simulation = cargo.simulation
    
    if request.method == 'POST':
        form = CargoForm(request.POST, instance=cargo)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cargo updated successfully!')
            return redirect('lng_planner:dashboard')
    else:
        form = CargoForm(instance=cargo)
    
    return render(request, 'lng_planner/cargo_form.html', {
        'form': form,
        'simulation': simulation,
        'title': 'Edit Cargo'
    })


@login_required
def delete_cargo(request, cargo_id):
    # Delete a cargo delivery from the user's simulation.
    cargo = get_object_or_404(Cargo, pk=cargo_id, simulation__user=request.user)
    cargo.delete()
    messages.success(request, 'Cargo deleted successfully!')
    return redirect('lng_planner:dashboard')


# ── Customer CRUD views ───────────────────────────────────────────────────
@login_required
def add_customer(request, simulation_id):
    # Add a new customer demand profile to the selected simulation.
    simulation = get_object_or_404(Simulation, pk=simulation_id, user=request.user)
    
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            customer = form.save(commit=False)
            customer.simulation = simulation
            customer.save()
            
            # DEBUG: Print all POST keys
            print("=" * 60)
            print("DEBUG - add_customer POST data:")
            for key, value in request.POST.items():
                if 'date_range' in key or 'daily' in key:
                    print(f"  {key} = {value}")
            print("=" * 60)
            
            # Handle additional date ranges (dynamic fields)
            date_ranges_saved = 0
            for key in request.POST:
                if key.startswith('date_ranges_') and key.endswith('_daily_demand'):
                    parts = key.split('_')
                    if len(parts) >= 4:
                        index = parts[2]  # date_ranges_1_daily_demand -> parts[2] = '1'
                        
                        supplier_id = request.POST.get(f'date_ranges_{index}_supplier')
                        daily_dem = request.POST.get(f'date_ranges_{index}_daily_demand')
                        fr_date = request.POST.get(f'date_ranges_{index}_from_date')
                        to_dt = request.POST.get(f'date_ranges_{index}_to_date')
                        
                        print(f"Processing range {index}: supplier={supplier_id}, demand={daily_dem}, from={fr_date}, to={to_dt}")
                        
                        if daily_dem and fr_date and to_dt:
                            # Get supplier if provided, ensure it belongs to same simulation
                            supplier = None
                            if supplier_id:
                                try:
                                    supplier = Supplier.objects.get(pk=supplier_id, simulation=simulation)
                                except Supplier.DoesNotExist:
                                    supplier = None  # Skip invalid suppliers
                            
                            CustomerDate.objects.create(
                                customer=customer,
                                supplier=supplier,
                                from_date=fr_date,
                                to_date=to_dt,
                                daily_demand=daily_dem
                            )
                            date_ranges_saved += 1
                            print(f"  ✓ Saved date range {index}")
            
            print(f"\nTotal date ranges saved: {date_ranges_saved}")
            print("=" * 60)
            
            messages.success(request, f'Customer added successfully! {date_ranges_saved} date range(s) saved.')
            return redirect('lng_planner:dashboard')
    else:
        form = CustomerForm()
    
    return render(request, 'lng_planner/customer_form.html', {
        'form': form,
        'simulation': simulation,
        'suppliers': Supplier.objects.filter(simulation=simulation),
        'title': 'Add Customer'
    })


@login_required
def edit_customer(request, customer_id):
    # Update a customer profile, including demand date ranges and priority.
    customer = get_object_or_404(Customer, pk=customer_id, simulation__user=request.user)
    simulation = customer.simulation
    
    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            
            # Handle primary date range updates
            from_date = request.POST.get('from_date')
            to_date = request.POST.get('to_date')
            daily_demand = request.POST.get('daily_demand')
            date_range_id = request.POST.get('date_range_id')
            supplier_id = request.POST.get('supplier')
            
            if from_date and to_date and daily_demand:
                # Get supplier if provided, ensure it belongs to same simulation
                supplier = None
                if supplier_id:
                    try:
                        supplier = Supplier.objects.get(pk=supplier_id, simulation=simulation)
                    except Supplier.DoesNotExist:
                        supplier = None
                
                if date_range_id:
                    # Update existing date range
                    date_range = get_object_or_404(CustomerDate, pk=date_range_id, customer=customer)
                    date_range.from_date = from_date
                    date_range.to_date = to_date
                    date_range.daily_demand = daily_demand
                    date_range.supplier = supplier
                    date_range.save()
                else:
                    # Create new date range
                    CustomerDate.objects.create(
                        customer=customer,
                        supplier=supplier,
                        from_date=from_date,
                        to_date=to_date,
                        daily_demand=daily_demand
                    )
            
            # Handle additional date ranges (dynamic fields)
            for key in request.POST:
                if key.startswith('date_ranges_') and key.endswith('_daily_demand'):
                    parts = key.split('_')
                    if len(parts) >= 4:
                        index = parts[2]  # date_ranges_1_daily_demand -> parts[2] = '1'
                        
                        supplier_id = request.POST.get(f'date_ranges_{index}_supplier')
                        daily_dem = request.POST.get(f'date_ranges_{index}_daily_demand')
                        fr_date = request.POST.get(f'date_ranges_{index}_from_date')
                        to_dt = request.POST.get(f'date_ranges_{index}_to_date')
                        
                        if daily_dem and fr_date and to_dt:
                            # Get supplier if provided, ensure it belongs to same simulation
                            supplier = None
                            if supplier_id:
                                try:
                                    supplier = Supplier.objects.get(pk=supplier_id, simulation=simulation)
                                except Supplier.DoesNotExist:
                                    supplier = None
                            
                            range_id = request.POST.get(f'date_ranges_{index}_date_range_id')
                            
                            if range_id:
                                date_range = get_object_or_404(CustomerDate, pk=range_id, customer=customer)
                                date_range.from_date = fr_date
                                date_range.to_date = to_dt
                                date_range.daily_demand = daily_dem
                                date_range.supplier = supplier
                                date_range.save()
                            else:
                                CustomerDate.objects.create(
                                    customer=customer,
                                    supplier=supplier,
                                    from_date=fr_date,
                                    to_date=to_dt,
                                    daily_demand=daily_dem
                                )
            
            messages.success(request, 'Customer updated successfully!')
            return redirect('lng_planner:dashboard')
    else:
        form = CustomerForm(instance=customer)
    
    return render(request, 'lng_planner/customer_form.html', {
        'form': form,
        'simulation': simulation,
        'suppliers': Supplier.objects.filter(simulation=simulation),
        'title': 'Edit Customer',
        'customer': customer
    })


@login_required
def delete_customer(request, customer_id):
    # Remove a customer and all its demand ranges from the simulation.
    customer = get_object_or_404(Customer, pk=customer_id, simulation__user=request.user)
    customer.delete()
    messages.success(request, 'Customer deleted successfully!')
    return redirect('lng_planner:dashboard')


@login_required
def add_customer_date_range(request, customer_id):
    # Add a new date range to an existing customer
    customer = get_object_or_404(Customer, pk=customer_id, simulation__user=request.user)
    
    if request.method == 'POST':
        form = CustomerDateForm(request.POST)
        if form.is_valid():
            date_range = form.save(commit=False)
            date_range.customer = customer
            date_range.save()
            messages.success(request, 'Date range added successfully!')
        else:
            messages.error(request, 'Error adding date range. Please check the form.')
    
    return redirect('lng_planner:dashboard')


@login_required
def edit_customer_date_range(request, date_range_id):
    # Edit an existing customer date range
    date_range = get_object_or_404(CustomerDate, pk=date_range_id, customer__simulation__user=request.user)
    
    if request.method == 'POST':
        form = CustomerDateForm(request.POST, instance=date_range)
        if form.is_valid():
            form.save()
            messages.success(request, 'Date range updated successfully!')
        else:
            messages.error(request, 'Error updating date range.')
    
    return redirect('lng_planner:dashboard')


@login_required
def delete_customer_date_range(request, date_range_id):
    # Delete a customer date range
    date_range = get_object_or_404(CustomerDate, pk=date_range_id, customer__simulation__user=request.user)
    customer_name = date_range.customer.name
    date_range.delete()
    messages.success(request, f'Date range for {customer_name} deleted successfully!')
    return redirect('lng_planner:dashboard')


# ── Refinery CRUD views ───────────────────────────────────────────────────
@login_required
def add_refinery(request, simulation_id):
    # Add a new refinery to the selected simulation.
    simulation = get_object_or_404(Simulation, pk=simulation_id, user=request.user)
    
    if request.method == 'POST':
        form = RefineryForm(request.POST)
        if form.is_valid():
            refinery = form.save(commit=False)
            refinery.simulation = simulation
            refinery.save()
            
            # DEBUG: Print all POST keys
            print("=" * 60)
            print("DEBUG - add_refinery POST data:")
            for key, value in request.POST.items():
                if 'date_range' in key or 'daily' in key:
                    print(f"  {key} = {value}")
            print("=" * 60)
            
            # Handle additional date ranges (dynamic fields) with supplier
            date_ranges_saved = 0
            for key in request.POST:
                if key.startswith('date_ranges_') and key.endswith('_daily_refinery_demand'):
                    parts = key.split('_')
                    if len(parts) >= 4:
                        index = parts[2]  # date_ranges_1_daily_refinery_demand -> parts[2] = '1'
                        
                        supplier_id = request.POST.get(f'date_ranges_{index}_supplier')
                        daily_dem = request.POST.get(f'date_ranges_{index}_daily_refinery_demand')
                        fr_date = request.POST.get(f'date_ranges_{index}_from_date')
                        to_dt = request.POST.get(f'date_ranges_{index}_to_date')
                        
                        print(f"Processing range {index}: supplier={supplier_id}, demand={daily_dem}, from={fr_date}, to={to_dt}")
                        
                        if daily_dem and fr_date and to_dt:
                            # Get supplier if provided, ensure it belongs to same simulation and plant
                            supplier = None
                            if supplier_id:
                                try:
                                    supplier = Supplier.objects.get(pk=supplier_id, simulation=simulation, plant=refinery.plant)
                                except Supplier.DoesNotExist:
                                    supplier = None  # Skip invalid suppliers
                            
                            RefineryDate.objects.create(
                                refinery=refinery,
                                supplier=supplier,
                                from_date=fr_date,
                                to_date=to_dt,
                                daily_refinery_demand=daily_dem
                            )
                            date_ranges_saved += 1
                            print(f"  ✓ Saved date range {index}")
            
            print(f"\nTotal date ranges saved: {date_ranges_saved}")
            print("=" * 60)
            
            messages.success(request, f'Refinery added successfully! {date_ranges_saved} date range(s) saved.')
            return redirect('lng_planner:dashboard')
    else:
        form = RefineryForm()
    
    # Get suppliers filtered by simulation for the form
    suppliers = Supplier.objects.filter(simulation=simulation)
    
    return render(request, 'lng_planner/refinery_form.html', {
        'form': form,
        'simulation': simulation,
        'suppliers': suppliers,
        'title': 'Add Refinery'
    })


@login_required
def edit_refinery(request, refinery_id):
    # Update a refinery profile, including demand date ranges and priority.
    refinery = get_object_or_404(Refinery, pk=refinery_id, simulation__user=request.user)
    simulation = refinery.simulation
    
    if request.method == 'POST':
        form = RefineryForm(request.POST, instance=refinery)
        if form.is_valid():
            form.save()
            
            # Handle primary date range updates
            from_date = request.POST.get('from_date')
            to_date = request.POST.get('to_date')
            daily_refinery_demand = request.POST.get('daily_refinery_demand')
            date_range_id = request.POST.get('date_range_id')
            
            if from_date and to_date and daily_refinery_demand:
                if date_range_id:
                    # Update existing date range
                    date_range = get_object_or_404(RefineryDate, pk=date_range_id, refinery=refinery)
                    date_range.from_date = from_date
                    date_range.to_date = to_date
                    date_range.daily_refinery_demand = daily_refinery_demand
                    date_range.save()
                else:
                    # Create new date range
                    RefineryDate.objects.create(
                        refinery=refinery,
                        from_date=from_date,
                        to_date=to_date,
                        daily_refinery_demand=daily_refinery_demand
                    )
            
            # Handle additional date ranges (dynamic fields)
            for key in request.POST:
                if key.startswith('date_ranges_') and key.endswith('_daily_refinery_demand'):
                    parts = key.split('_')
                    if len(parts) >= 4:
                        index = parts[2]  # date_ranges_1_daily_refinery_demand -> parts[2] = '1'
                        
                        daily_dem = request.POST.get(f'date_ranges_{index}_daily_refinery_demand')
                        fr_date = request.POST.get(f'date_ranges_{index}_from_date')
                        to_dt = request.POST.get(f'date_ranges_{index}_to_date')
                        
                        if daily_dem and fr_date and to_dt:
                            range_id = request.POST.get(f'date_ranges_{index}_date_range_id')
                            
                            if range_id:
                                date_range = get_object_or_404(RefineryDate, pk=range_id, refinery=refinery)
                                date_range.from_date = fr_date
                                date_range.to_date = to_dt
                                date_range.daily_refinery_demand = daily_dem
                                date_range.save()
                            else:
                                RefineryDate.objects.create(
                                    refinery=refinery,
                                    from_date=fr_date,
                                    to_date=to_dt,
                                    daily_refinery_demand=daily_dem
                                )
            
            messages.success(request, 'Refinery updated successfully!')
            return redirect('lng_planner:dashboard')
    else:
        form = RefineryForm(instance=refinery)
    
    return render(request, 'lng_planner/refinery_form.html', {
        'form': form,
        'simulation': simulation,
        'title': 'Edit Refinery',
        'refinery': refinery
    })


@login_required
def delete_refinery(request, refinery_id):
    # Remove a refinery and all its demand ranges from the simulation.
    refinery = get_object_or_404(Refinery, pk=refinery_id, simulation__user=request.user)
    refinery.delete()
    messages.success(request, 'Refinery deleted successfully!')
    return redirect('lng_planner:dashboard')


@login_required
def add_refinery_date_range(request, refinery_id):
    # Add a new date range to an existing refinery
    refinery = get_object_or_404(Refinery, pk=refinery_id, simulation__user=request.user)
    
    if request.method == 'POST':
        form = RefineryDateForm(request.POST)
        if form.is_valid():
            date_range = form.save(commit=False)
            date_range.refinery = refinery
            date_range.save()
            messages.success(request, 'Date range added successfully!')
        else:
            messages.error(request, 'Error adding date range. Please check the form.')
    
    return redirect('lng_planner:dashboard')


@login_required
def edit_refinery_date_range(request, date_range_id):
    # Edit an existing refinery date range
    date_range = get_object_or_404(RefineryDate, pk=date_range_id, refinery__simulation__user=request.user)
    
    if request.method == 'POST':
        form = RefineryDateForm(request.POST, instance=date_range)
        if form.is_valid():
            form.save()
            messages.success(request, 'Date range updated successfully!')
        else:
            messages.error(request, 'Error updating date range.')
    
    return redirect('lng_planner:dashboard')


@login_required
def delete_refinery_date_range(request, date_range_id):
    # Delete a refinery date range
    date_range = get_object_or_404(RefineryDate, pk=date_range_id, refinery__simulation__user=request.user)
    refinery_name = date_range.refinery.name
    date_range.delete()
    messages.success(request, f'Date range for {refinery_name} deleted successfully!')
    return redirect('lng_planner:dashboard')


# ── Simulation Comment views ────────────────────────────────────────────────
@login_required
def add_comment(request, simulation_id):
    # Add a comment to the simulation
    simulation = get_object_or_404(Simulation, pk=simulation_id, user=request.user)
    
    if request.method == 'POST':
        comment_text = request.POST.get('comment')
        if comment_text:
            SimulationComment.objects.create(
                simulation=simulation,
                comment=comment_text,
                created_by=request.user
            )
            messages.success(request, 'Comment added successfully!')
        else:
            messages.error(request, 'Comment cannot be empty.')
    
    return redirect('lng_planner:dashboard')


@login_required
def edit_comment(request, comment_id):
    # Edit an existing comment
    comment = get_object_or_404(SimulationComment, pk=comment_id, simulation__user=request.user)
    
    # Only allow author or planning admin to edit
    if comment.created_by != request.user and not is_planning_admin(request.user):
        messages.error(request, 'You do not have permission to edit this comment.')
        return redirect('lng_planner:dashboard')
    
    if request.method == 'POST':
        comment_text = request.POST.get('comment')
        if comment_text:
            comment.comment = comment_text
            comment.save()
            messages.success(request, 'Comment updated successfully!')
        else:
            messages.error(request, 'Comment cannot be empty.')
    
    return redirect('lng_planner:dashboard')


@login_required
def delete_comment(request, comment_id):
    # Delete a comment
    comment = get_object_or_404(SimulationComment, pk=comment_id, simulation__user=request.user)
    
    # Only allow author or planning admin to delete
    if comment.created_by != request.user and not is_planning_admin(request.user):
        messages.error(request, 'You do not have permission to delete this comment.')
        return redirect('lng_planner:dashboard')
    
    comment.delete()
    messages.success(request, 'Comment deleted successfully!')
    return redirect('lng_planner:dashboard')


# ── Export / import views ─────────────────────────────────────────────────
@login_required
def export_excel(request, simulation_id):
    """Export simulation data to Excel with colour coding matching the dashboard."""
    from datetime import date as date_type
    from openpyxl.styles import Border, Side
    
    # Get filter parameters from GET request
    from datetime import datetime
    
    def dates_overlap(start1, end1, start2, end2):
        """Check if two date ranges overlap"""
        return start1 <= end2 and start2 <= end1
    
    def entity_has_active_ranges(entity_date_ranges, filter_from, filter_to):
        """Check if an entity has at least one date range overlapping with filter period"""
        if not filter_from or not filter_to:
            return True  # No filter = show all
        
        for dr in entity_date_ranges.all():
            if dates_overlap(dr.from_date, dr.to_date, filter_from, filter_to):
                return True
        return False
    
    def cargo_in_range(cargo_delivery_date, filter_from, filter_to):
        """Check if cargo delivery date falls within filter period"""
        if not filter_from or not filter_to:
            return True  # No filter = show all
        return filter_from <= cargo_delivery_date <= filter_to
    
    # Handle date range filter from GET parameters
    filter_from_date = None
    filter_to_date = None
    
    if request.GET.get('from_date') and request.GET.get('to_date'):
        try:
            filter_from_date = datetime.strptime(request.GET['from_date'], '%Y-%m-%d').date()
            filter_to_date = datetime.strptime(request.GET['to_date'], '%Y-%m-%d').date()
        except ValueError:
            pass  # Invalid date format, ignore filter
    
    # Get the simulation being exported
    simulation = get_object_or_404(Simulation, pk=simulation_id)
    
    # Debug: Log which simulation is being exported
    print(f"\n📊 EXPORTING SIMULATION:")
    print(f"  Simulation ID: {simulation.id}")
    print(f"  Simulation Name: {simulation.name}")
    print(f"  Is Master: {simulation.is_master}")
    print(f"  User: {simulation.user.username if simulation.user else 'N/A (Master)'}")
    print(f"  Request User: {request.user.username}")
    print("="*60 + "\n")
    
    # Check permissions - allow exporting own simulations or master
    if not simulation.is_master and simulation.user != request.user:
        messages.error(request, 'You do not have permission to export this simulation')
        return redirect('lng_planner:dashboard')
    
    # Calculate daily data and apply date filter
    daily_data = calculate_daily_data(simulation)
    
    # Apply date filter to daily_data if filter is active
    if filter_from_date and filter_to_date:
        daily_data = [day for day in daily_data 
                     if filter_from_date <= day['date'] <= filter_to_date]
    
    plants = Plant.objects.filter(
        id__in=simulation.plant_inventories.values_list('plant_id', flat=True)
    )
    today = date_type.today()
    
    # Build summary dictionaries for customers, refineries, and suppliers (same logic as get_simulation_data)
    customers_summary = {}
    for plant in plants:
        summary = {}
        qs = simulation.customers.filter(plant=plant).order_by('name')
        for c in qs:
            # Skip customer if no date ranges overlap with filter period
            if not entity_has_active_ranges(c.date_ranges, filter_from_date, filter_to_date):
                continue
                
            range_list = []
            for cd in c.date_ranges.all():
                # Only include date ranges that overlap with filter period
                if filter_from_date and filter_to_date:
                    if not dates_overlap(cd.from_date, cd.to_date, filter_from_date, filter_to_date):
                        continue
                
                range_dict = {
                    'id': cd.id,
                    'from_date': cd.from_date,
                    'to_date': cd.to_date,
                    'daily_demand': float(cd.daily_demand),
                }
                range_list.append(range_dict)
            
            # Only add customer if it has at least one valid date range
            if range_list:
                if c.name in summary:
                    summary[c.name]['ranges'].extend(range_list)
                else:
                    summary[c.name] = {
                        'id': c.id,
                        'name': c.name,
                        'plant': plant,
                        'preference': c.preference,
                        'ranges': range_list
                    }
        customers_summary[plant.id] = list(summary.values())

    refineries_summary = {}
    for plant in plants:
        summary = {}
        qs = simulation.refineries.filter(plant=plant).order_by('name')
        for r in qs:
            # Skip refinery if no date ranges overlap with filter period
            if not entity_has_active_ranges(r.date_ranges, filter_from_date, filter_to_date):
                continue
                
            range_list = []
            for rd in r.date_ranges.all():
                # Only include date ranges that overlap with filter period
                if filter_from_date and filter_to_date:
                    if not dates_overlap(rd.from_date, rd.to_date, filter_from_date, filter_to_date):
                        continue
                
                range_dict = {
                    'id': rd.id,
                    'from_date': rd.from_date,
                    'to_date': rd.to_date,
                    'daily_refinery_demand': float(rd.daily_refinery_demand),
                }
                range_list.append(range_dict)
            
            # Only add refinery if it has at least one valid date range
            if range_list:
                if r.name in summary:
                    summary[r.name]['ranges'].extend(range_list)
                else:
                    summary[r.name] = {
                        'id': r.id,
                        'name': r.name,
                        'plant': plant,
                        'preference': r.preference,
                        'ranges': range_list
                    }
        refineries_summary[plant.id] = list(summary.values())

    suppliers_summary = {}
    for plant in plants:
        summary = {}
        qs = simulation.suppliers.filter(plant=plant).order_by('name')
        for s in qs:
            # Skip supplier if no date ranges overlap with filter period
            if not entity_has_active_ranges(s.date_ranges, filter_from_date, filter_to_date):
                continue
                
            range_list = []
            for sd in s.date_ranges.all():
                # Only include date ranges that overlap with filter period
                if filter_from_date and filter_to_date:
                    if not dates_overlap(sd.from_date, sd.to_date, filter_from_date, filter_to_date):
                        continue
                
                range_dict = {
                    'id': sd.id,
                    'from_date': sd.from_date,
                    'to_date': sd.to_date,
                    'daily_supply': float(sd.daily_supply),
                }
                range_list.append(range_dict)
            
            # Only add supplier if it has at least one valid date range
            if range_list:
                if s.name in summary:
                    summary[s.name]['ranges'].extend(range_list)
                else:
                    summary[s.name] = {
                        'id': s.id,
                        'name': s.name,
                        'plant': plant,
                        'preference': s.preference,
                        'ranges': range_list
                    }
        suppliers_summary[plant.id] = list(summary.values())

    # ── Colour palette (matches Tailwind classes used in dashboard) ────────────
    # Header row  → gray-300
    C_HEADER        = 'FFD1D5DB'
    # Plant title  → blue-100
    C_PLANT_TITLE   = 'FFDBEAFE'
    # Supplier rows → green-50
    C_SUPPLIER      = 'FFF0FDF4'
    # Cargo rows   → teal-50
    C_CARGO         = 'FFF0FDFA'
    # Customer rows → orange-50
    C_CUSTOMER      = 'FFFFF7ED'
    # Refinery rows → indigo-50
    C_REFINERY      = 'FFEEF2FD'
    # Total supply  → green-100
    C_TOT_SUPPLY    = 'FFDCFCE7'
    # Total demand  → orange-100
    C_TOT_DEMAND    = 'FFFEE3C8' # approx orange-100
    # Closing inv   → blue-100
    C_CLOSING_INV   = 'FFDBEAFE'
    # Negative inv  → red-100
    C_NEGATIVE      = 'FFFEE2E2'
    # Overall totals header → gray-200
    C_GRAND_TITLE   = 'FFE5E7EB'
    # Overall supply → green-200
    C_GRAND_SUPPLY  = 'FFBBF7D0'
    # Overall demand → orange-200
    C_GRAND_DEMAND  = 'FFFED7AA'
    # Overall inv   → blue-200
    C_GRAND_INV     = 'FFBFDBFE'
    # Spacer / blank
    C_WHITE         = 'FFFFFFFF'

    # Font colours
    FC_GREEN        = 'FF166534'   # green-700 text
    FC_TEAL         = 'FF0F766E'   # teal-700
    FC_ORANGE       = 'FF9A3412'   # orange-700
    FC_INDIGO       = 'FF5B2E8C'   # indigo-700
    FC_BLUE         = 'FF1D4ED8'   # blue-700
    FC_RED          = 'FFB91C1C'   # red-700
    FC_BLUE_HEADER  = 'FF1E3A5F'   # dark blue for plant title
    FC_GRAY         = 'FF374151'   # gray-700

    def _fill(hex_color):
        return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

    def _font(bold=False, color=None, size=11):
        return Font(bold=bold, color=color or 'FF000000', size=size)

    def _center():
        return Alignment(horizontal='center', vertical='center', wrap_text=False)

    def _left():
        return Alignment(horizontal='left', vertical='center')

    thin_side = Side(style='thin', color='FFD1D5DB')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def _style_row(ws_row, bg, fc=None, bold=False, first_left=True):
        """Apply fill, font, border and alignment to every cell in ws_row."""
        for i, cell in enumerate(ws_row):
            cell.fill = _fill(bg)
            cell.font = _font(bold=bold, color=fc)
            cell.border = thin_border
            cell.alignment = _left() if (i == 0 and first_left) else _center()

    # ── Workbook setup ─────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'LNG Planning'
    ws.freeze_panes = 'B2'          # freeze header row + item column

    # Column widths
    ws.column_dimensions['A'].width = 38
    for col_idx in range(2, len(daily_data) + 2):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = 14

    # ── Header row (dates) ────────────────────────────────────────────────────
    header = ['Item'] + [day['date'].strftime('%b %d, %Y') for day in daily_data]
    ws.append(header)
    _style_row(ws[ws.max_row], C_HEADER, fc=FC_GRAY, bold=True)
    # Mark today's column with a slightly darker shade so it stands out
    for col_idx, day in enumerate(daily_data, start=2):
        if day['date'] == today:
            ws.cell(row=1, column=col_idx).fill = _fill('FFADB5BD')  # gray-400

    # ── Per-plant rows ────────────────────────────────────────────────────────
    for plant in plants:
        # Plant title row  →  🏭 Plant Name
        ws.append([f'🏭  {plant.name}'] + ['' for _ in daily_data])
        plant_title_row = ws[ws.max_row]
        for cell in plant_title_row:
            cell.fill = _fill(C_PLANT_TITLE)
            cell.font = _font(bold=True, color=FC_BLUE_HEADER, size=12)
            cell.border = thin_border
            cell.alignment = _left()

        # ── Suppliers ────────────────────────────────────────────────────────
        for supplier in simulation.suppliers.filter(plant=plant):
            # Skip supplier if no date ranges overlap with filter period
            if not entity_has_active_ranges(supplier.date_ranges, filter_from_date, filter_to_date):
                continue
                
            row = [f'    {supplier.name}']
            for day in daily_data:
                pd = day['plant_data'].get(plant.id, {})
                # Sum ALL active date ranges for this supplier (not just first match)
                total_supply = sum(
                    s['amount'] for s in pd.get('supplies', [])
                    if s['type'] == 'supplier' and s['name'] == supplier.name
                )
                row.append(total_supply if total_supply > 0 else '')
            ws.append(row)
            _style_row(ws[ws.max_row], C_SUPPLIER, fc=FC_GREEN)

        # ── Cargos ───────────────────────────────────────────────────────────
        for cargo in simulation.cargos.filter(plant=plant):
            # Skip cargo if delivery date is outside filter range
            if not cargo_in_range(cargo.delivery_date, filter_from_date, filter_to_date):
                continue
                
            row = [f'    {cargo.cargo_name}']
            for day in daily_data:
                pd = day['plant_data'].get(plant.id, {})
                # Sum ALL active date ranges for this cargo (not just first match)
                total_cargo = sum(
                    s['amount'] for s in pd.get('supplies', [])
                    if s['type'] == 'cargo' and s['name'] == cargo.cargo_name
                )
                row.append(total_cargo if total_cargo > 0 else '')
            ws.append(row)
            _style_row(ws[ws.max_row], C_CARGO, fc=FC_TEAL)

        # ── Total Supply ─────────────────────────────────────────────────────
        row = [f'    Total Supply — {plant.name}']
        for day in daily_data:
            pd = day['plant_data'].get(plant.id, {})
            supply = pd.get('total_supplier_supply', 0)
            row.append(supply if supply > 0 else '')
        ws.append(row)
        _style_row(ws[ws.max_row], C_TOT_SUPPLY, fc=FC_GREEN, bold=True)

        # ── Customers ────────────────────────────────────────────────────────
        for customer in customers_summary.get(plant.id, []):
            customer_name = customer['name']
            row = [f'    {customer_name}']
            for day in daily_data:
                pd = day['plant_data'].get(plant.id, {})
                val = ''
                if pd and pd.get('customers_today'):
                    cust_data = pd['customers_today'].get(customer_name)
                    if cust_data:
                        val = f"{cust_data['total_demand']:.0f}"
                row.append(val)
            ws.append(row)
            _style_row(ws[ws.max_row], C_CUSTOMER, fc=FC_ORANGE)

        # ── Total Customer Demand ────────────────────────────────────────────
        row = [f'    Total Customer Demand — {plant.name}']
        for day in daily_data:
            pd = day['plant_data'].get(plant.id, {})
            demand = pd.get('total_customer_demand', 0)
            row.append(demand if demand > 0 else '')
        ws.append(row)
        _style_row(ws[ws.max_row], C_TOT_DEMAND, fc=FC_ORANGE, bold=True)

        # ── Total Refinery Demand ────────────────────────────────────────────
        row = [f'    Total Refinery Demand — {plant.name}']
        for day in daily_data:
            pd = day['plant_data'].get(plant.id, {})
            demand = pd.get('total_refinery_demand', 0)
            row.append(demand if demand > 0 else '')
        ws.append(row)
        _style_row(ws[ws.max_row], C_REFINERY, fc=FC_INDIGO, bold=True)

        # ── Refineries ───────────────────────────────────────────────────────
        for refinery in refineries_summary.get(plant.id, []):
            refinery_name = refinery['name']
            row = [f'    {refinery_name}']
            for day in daily_data:
                pd = day['plant_data'].get(plant.id, {})
                val = ''
                if pd and pd.get('refineries_today'):
                    ref_data = pd['refineries_today'].get(refinery_name)
                    if ref_data:
                        val = f"{ref_data['total_demand']:.0f}"
                row.append(val)
            ws.append(row)
            _style_row(ws[ws.max_row], C_REFINERY, fc=FC_INDIGO)

        # ── Plant Inventory ───────────────────────────────────────────────────
        row = [f'    Plant Inventory — {plant.name}']
        for day in daily_data:
            pd = day['plant_data'].get(plant.id, {})
            plant_inv = pd.get('plant_inventory', 0)
            row.append(plant_inv if plant_inv > 0 else '')
        ws.append(row)
        _style_row(ws[ws.max_row], C_CLOSING_INV, fc=FC_BLUE, bold=True)

        # ── Cargo Inventory ───────────────────────────────────────────────────
        row = [f'    Cargo Inventory — {plant.name}']
        for day in daily_data:
            pd = day['plant_data'].get(plant.id, {})
            cargo_inv = pd.get('cargo_inventory', 0)
            if cargo_inv < 0:
                row.append(f"{cargo_inv:.0f}")
            elif cargo_inv > 0:
                row.append(cargo_inv)
            else:
                row.append('')
        ws.append(row)
        cargo_row = ws[ws.max_row]
        for col_idx, day in enumerate(daily_data, start=2):
            pd = day['plant_data'].get(plant.id, {})
            cell = cargo_row[col_idx - 1]
            cargo_inv = pd.get('cargo_inventory', 0)
            if cargo_inv < 0:
                cell.fill = _fill(C_NEGATIVE)
                cell.font = _font(bold=True, color=FC_RED)
            elif cargo_inv > 0:
                cell.fill = _fill(C_CLOSING_INV)
                cell.font = _font(bold=True, color=FC_BLUE)
            else:
                cell.fill = _fill(C_WHITE)
                cell.font = _font(color=FC_GRAY)
            cell.border = thin_border
            cell.alignment = _center()
        # Style the label cell
        cargo_row[0].fill = _fill(C_CLOSING_INV)
        cargo_row[0].font = _font(bold=True, color=FC_BLUE)
        cargo_row[0].border = thin_border
        cargo_row[0].alignment = _left()

        # ── Closing Inventory ─────────────────────────────────────────────────
        row_data = [f'    Closing Inventory — {plant.name}']
        for day in daily_data:
            pd = day['plant_data'].get(plant.id, {})
            closing_inventory = pd.get('closing_inventory', pd.get('inventory', 0))
            backlog = pd.get('backlog', 0)
            if backlog > 0:
                row_data.append(f"{closing_inventory:.0f} ({backlog:.0f})")
            else:
                row_data.append(closing_inventory)
        ws.append(row_data)
        inv_row = ws[ws.max_row]
        for col_idx, day in enumerate(daily_data, start=2):
            pd = day['plant_data'].get(plant.id, {})
            cell = inv_row[col_idx - 1]
            is_neg = pd.get('is_negative', False)
            cell.fill = _fill(C_NEGATIVE if is_neg else C_CLOSING_INV)
            cell.font = _font(bold=True, color=FC_RED if is_neg else FC_BLUE)
            cell.border = thin_border
            cell.alignment = _center()
        # Style the label cell
        inv_row[0].fill = _fill(C_CLOSING_INV)
        inv_row[0].font = _font(bold=True, color=FC_BLUE)
        inv_row[0].border = thin_border
        inv_row[0].alignment = _left()

        # ── Projected Inventory Balance (Uncapped) ────────────────────────────
        row_data = [f'    Projected Balance — {plant.name}']
        for day in daily_data:
            pd = day['plant_data'].get(plant.id, {})
            projected_balance = pd.get('projected_balance', 0)
            if projected_balance < 0:
                row_data.append(f"-{abs(projected_balance):.0f}")
            elif projected_balance > 0:
                row_data.append(f"{projected_balance:.0f}")
            else:
                row_data.append("")
        ws.append(row_data)
        proj_row = ws[ws.max_row]
        for col_idx, day in enumerate(daily_data, start=2):
            pd = day['plant_data'].get(plant.id, {})
            cell = proj_row[col_idx - 1]
            projected_balance = pd.get('projected_balance', 0)
            if projected_balance < 0:
                cell.fill = _fill(C_NEGATIVE)
                cell.font = _font(bold=True, color=FC_RED)
            elif projected_balance > 0:
                cell.fill = _fill(C_CLOSING_INV)
                cell.font = _font(bold=True, color=FC_GREEN)
            else:
                cell.fill = _fill(C_CLOSING_INV)
                cell.font = _font(bold=False, color=FC_GRAY)
            cell.border = thin_border
            cell.alignment = _center()
        proj_row[0].fill = _fill(C_CLOSING_INV)
        proj_row[0].font = _font(bold=True, color=FC_BLUE)
        proj_row[0].border = thin_border
        proj_row[0].alignment = _left()

        # Blank spacer row between plants
        ws.append([''] + ['' for _ in daily_data])
        _style_row(ws[ws.max_row], C_WHITE)

    # ── Overall Totals section ────────────────────────────────────────────────
    ws.append(['📊  TOTAL (All Plants)'] + ['' for _ in daily_data])
    _style_row(ws[ws.max_row], C_GRAND_TITLE, fc=FC_GRAY, bold=True)

    row = ['    Total Supply (All Plants)']
    for day in daily_data:
        row.append(day['total_supply'] if day['total_supply'] > 0 else '')
    ws.append(row)
    _style_row(ws[ws.max_row], C_GRAND_SUPPLY, fc=FC_GREEN, bold=True)

    row = ['    Total Customer Demand (All Plants)']
    for day in daily_data:
        demand = day.get('total_customer_demand', 0)
        row.append(demand if demand > 0 else '')
    ws.append(row)
    _style_row(ws[ws.max_row], C_GRAND_DEMAND, fc=FC_ORANGE, bold=True)

    row = ['    Total Refinery Demand (All Plants)']
    for day in daily_data:
        demand = day.get('total_refinery_demand', 0)
        row.append(demand if demand > 0 else '')
    ws.append(row)
    _style_row(ws[ws.max_row], C_GRAND_DEMAND, fc=FC_INDIGO, bold=True)

    row = ['    Total Closing Inventory (All Plants)']
    for day in daily_data:
        total_inventory = day['total_inventory']
        if total_inventory < 0:
            row.append(f"0 ({abs(total_inventory):.0f})")
        else:
            row.append(total_inventory)
    ws.append(row)
    _style_row(ws[ws.max_row], C_GRAND_INV, fc=FC_BLUE, bold=True)

    row = ['    Total Projected Balance (All Plants)']
    for day in daily_data:
        total_projected_balance = day.get('total_projected_balance', 0)
        if total_projected_balance < 0:
            row.append(f"-{abs(total_projected_balance):.0f}")
        elif total_projected_balance > 0:
            row.append(f"{total_projected_balance:.0f}")
        else:
            row.append("")
    ws.append(row)
    _style_row(ws[ws.max_row], C_GRAND_INV, fc=FC_BLUE, bold=True)

    # ── Comments Section ─────────────────────────────────────────────────────
    ws.append([])  # Blank spacer row
    ws.append(['💬 SIMULATION COMMENTS'])
    comment_header = ws[ws.max_row]
    for cell in comment_header:
        cell.fill = _fill(C_HEADER)
        cell.font = _font(bold=True, color=FC_GRAY)
        cell.border = thin_border
        cell.alignment = _left()

    comments = simulation.comments.select_related('created_by').order_by('-created_at')
    if comments:
        for comment in comments:
            author_name = comment.created_by.get_full_name() or comment.created_by.username
            timestamp = comment.created_at.strftime('%b %d, %Y %I:%M %p')
            ws.append([f"{author_name} ({timestamp})", comment.comment])
            comment_row = ws[ws.max_row]
            for cell in comment_row:
                cell.border = thin_border
                cell.alignment = _left()
    else:
        ws.append(['No comments yet.'])
        no_comment_row = ws[ws.max_row]
        for cell in no_comment_row:
            cell.border = thin_border
            cell.alignment = _left()

    # ── HTTP response ─────────────────────────────────────────────────────────
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename=LNG_Planning_{simulation.start_date}_to_{simulation.end_date}.xlsx'
    )
    wb.save(response)
    return response


@login_required
def export_json(request, simulation_id):
    # Export the current simulation and associated entities as JSON.
    """Export simulation as JSON"""
    simulation = get_object_or_404(Simulation, pk=simulation_id, user=request.user)
    
    data = {
        'simulation': {
            'name': simulation.name,
            'start_date': str(simulation.start_date),
            'end_date': str(simulation.end_date),
        },
        'suppliers': [
            {
                'name': s.name,
                'plant_id': s.plant_id,
                'preference': s.preference,
                'date_ranges': [
                    {
                        'from_date': str(sd.from_date),
                        'to_date': str(sd.to_date),
                        'daily_supply': float(sd.daily_supply)
                    }
                    for sd in s.date_ranges.all()
                ]
            }
            for s in simulation.suppliers.all()
        ],
        'cargos': [
            {
                'cargo_name': c.cargo_name,
                'plant_id': c.plant_id,
                'date': str(c.delivery_date),
                'amount': float(c.amount)
            }
            for c in simulation.cargos.all()
        ],
        'customers': [
            {
                'name': c.name,
                'plant_id': c.plant_id,
                'preference': c.preference,
                'date_ranges': [
                    {
                        'from_date': str(cd.from_date),
                        'to_date': str(cd.to_date),
                        'daily_demand': float(cd.daily_demand)
                    }
                    for cd in c.date_ranges.all()
                ]
            }
            for c in simulation.customers.all()
        ],
        'refineries': [
            {
                'name': r.name,
                'plant_id': r.plant_id,
                'preference': r.preference,
                'date_ranges': [
                    {
                        'from_date': str(rd.from_date),
                        'to_date': str(rd.to_date),
                        'daily_refinery_demand': float(rd.daily_refinery_demand)
                    }
                    for rd in r.date_ranges.all()
                ]
            }
            for r in simulation.refineries.all()
        ],
        'plant_inventories': [
            {
                'plant_id': pi.plant_id,
                'opening_inventory': float(pi.opening_inventory)
            }
            for pi in simulation.plant_inventories.all()
        ]
    }
    
    response = JsonResponse(data)
    response['Content-Disposition'] = f'attachment; filename=LNG_Simulation_{simulation.name}.json'
    return response


@login_required
def import_json(request, simulation_id):
    # Import suppliers, cargos, customers, and refineries from provided JSON data.
    """Import data from JSON"""
    simulation = get_object_or_404(Simulation, pk=simulation_id, user=request.user)
    
    if request.method == 'POST':
        form = JSONUploadForm(request.POST, request.FILES)
        if form.is_valid():
            json_file = request.FILES['json_file']
            try:
                data = json.loads(json_file.read().decode('utf-8'))
                
                # Clear existing simulation data before importing new records.
                simulation.suppliers.all().delete()
                simulation.cargos.all().delete()
                simulation.customers.all().delete()
                simulation.refineries.all().delete()
                
                if 'suppliers' in data:
                    for s in data['suppliers']:
                        supplier = Supplier.objects.create(
                            simulation=simulation,
                            name=s['name'],
                            plant_id=s.get('plant_id'),
                            preference=s.get('preference', 1)
                        )
                        # Create date ranges
                        for dr in s.get('date_ranges', []):
                            SupplierDate.objects.create(
                                supplier=supplier,
                                from_date=dr['from_date'],
                                to_date=dr['to_date'],
                                daily_supply=dr['daily_supply']
                            )
                
                if 'cargos' in data:
                    for c in data['cargos']:
                        Cargo.objects.create(
                            simulation=simulation,
                            cargo_name=c['cargo_name'],
                            delivery_date=c['date'],
                            amount=c['amount']
                        )
                
                if 'customers' in data:
                    for c in data['customers']:
                        customer = Customer.objects.create(
                            simulation=simulation,
                            name=c['name'],
                            plant_id=c.get('plant_id'),
                            preference=c.get('preference', 1)
                        )
                        # Create date ranges
                        for dr in c.get('date_ranges', []):
                            CustomerDate.objects.create(
                                customer=customer,
                                from_date=dr['from_date'],
                                to_date=dr['to_date'],
                                daily_demand=dr['daily_demand']
                            )
                
                if 'refineries' in data:
                    for r in data['refineries']:
                        refinery = Refinery.objects.create(
                            simulation=simulation,
                            name=r['name'],
                            plant_id=r.get('plant_id'),
                            preference=r.get('preference', 1)
                        )
                        # Create date ranges
                        for dr in r.get('date_ranges', []):
                            RefineryDate.objects.create(
                                refinery=refinery,
                                from_date=dr['from_date'],
                                to_date=dr['to_date'],
                                daily_refinery_demand=dr['daily_refinery_demand']
                            )
                
                if 'plant_inventories' in data:
                    for pi in data['plant_inventories']:
                        PlantInventory.objects.update_or_create(
                            simulation=simulation,
                            plant_id=pi['plant_id'],
                            defaults={'opening_inventory': pi['opening_inventory']}
                        )
                
                messages.success(request, 'Data imported successfully!')
                return redirect('lng_planner:dashboard')
                
            except Exception as e:
                messages.error(request, f'Error importing JSON: {str(e)}')
    else:
        form = JSONUploadForm()
    
    return render(request, 'lng_planner/import_json.html', {
        'form': form,
        'simulation': simulation
    })


# HELPER FUNCTIONS

def get_simulation_data(simulation, filter_from_date=None, filter_to_date=None):
    # Prepare all data required by the dashboard for a given simulation.
    # This includes daily projections, plant stats, customer summaries, and alerts.
    from datetime import date
    from django.db.models import Q
    
    # OPTIMIZATION: Pre-calculate filtered entity IDs at database level instead of Python loops
    
    # Calculate daily data first (this will be filtered later)
    daily_data = calculate_daily_data(simulation)
    
    # Apply date filter to daily_data if filter is active
    if filter_from_date and filter_to_date:
        daily_data = [day for day in daily_data 
                     if filter_from_date <= day['date'] <= filter_to_date]
    
    plants = Plant.objects.filter(
        id__in=simulation.plant_inventories.values_list('plant_id', flat=True)
    )
    
    # OPTIMIZATION: Pre-filter entities using database-level queries
    
    # Get filtered supplier IDs (suppliers with at least one date range overlapping filter period)
    if filter_from_date and filter_to_date:
        filtered_supplier_ids = simulation.suppliers.filter(
            Q(date_ranges__from_date__lte=filter_to_date) & 
            Q(date_ranges__to_date__gte=filter_from_date)
        ).values_list('id', flat=True).distinct()
        
        filtered_customer_ids = simulation.customers.filter(
            Q(date_ranges__from_date__lte=filter_to_date) & 
            Q(date_ranges__to_date__gte=filter_from_date)
        ).values_list('id', flat=True).distinct()
        
        filtered_refinery_ids = simulation.refineries.filter(
            Q(date_ranges__from_date__lte=filter_to_date) & 
            Q(date_ranges__to_date__gte=filter_from_date)
        ).values_list('id', flat=True).distinct()
        
        filtered_cargo_ids = simulation.cargos.filter(
            delivery_date__gte=filter_from_date,
            delivery_date__lte=filter_to_date
        ).values_list('id', flat=True)
    else:
        # No filter - get all IDs
        filtered_supplier_ids = simulation.suppliers.values_list('id', flat=True)
        filtered_customer_ids = simulation.customers.values_list('id', flat=True)
        filtered_refinery_ids = simulation.refineries.values_list('id', flat=True)
        filtered_cargo_ids = simulation.cargos.values_list('id', flat=True)
    
    # Find first negative date for each plant so dashboard can highlight inventory issues.
    plant_alerts = {}
    for plant in plants:
        first_negative = next(
            (d for d in daily_data if d['plant_data'].get(plant.id, {}).get('is_negative', False)),
            None
        )
        if first_negative:
            plant_alerts[plant.id] = {
                'plant': plant,
                'date': first_negative['date'],
                'inventory': first_negative['plant_data'][plant.id]['inventory']
            }
    
    # Calculate totals split between historical and future days based on today.
    today = date.today()
    
    # ── Total statistics ──────────────────────────────────────────────────────
    total_opening_inventory = sum(
        pi.opening_inventory for pi in simulation.plant_inventories.all()
    )

    total_supplied_till_today = 0
    total_demand_till_today = 0
    total_current_inventory = 0
    total_upcoming_supply = 0
    total_upcoming_demand = 0

    for day in daily_data:
        if day['date'] <= today:
            total_supplied_till_today += day['total_supply']
            total_demand_till_today   += day['total_demand']
            total_current_inventory    = day['total_inventory']   # last value = today
        else:
            total_upcoming_supply += day['total_supply']
            total_upcoming_demand += day['total_demand']

    # ── Plant-wise statistics ─────────────────────────────────────────────────
    plant_stats = {}
    for plant in plants:
        plant_inventory = simulation.plant_inventories.filter(plant=plant).first()
        opening_inv = float(plant_inventory.opening_inventory) if plant_inventory else 0

        stats = {
            'suppliers':              simulation.suppliers.filter(plant=plant, id__in=filtered_supplier_ids).count(),
            'customers':              simulation.customers.filter(plant=plant, id__in=filtered_customer_ids).count(),
            'cargos':                 simulation.cargos.filter(plant=plant, id__in=filtered_cargo_ids).count(),
            'opening_inventory':      opening_inv,
            'supplied_qty_till_today': 0,
            'received_qty_till_today': 0,
            'current_inventory':       0,
            'upcoming_supply':         0,
            'upcoming_demand':         0,
        }

        for day in daily_data:
            plant_data = day['plant_data'].get(plant.id, {})
            if day['date'] <= today:
                stats['supplied_qty_till_today'] += plant_data.get('supply', 0)
                stats['received_qty_till_today'] += plant_data.get('demand', 0)
                stats['current_inventory']        = plant_data.get('inventory', 0)
            else:
                stats['upcoming_supply'] += plant_data.get('supply', 0)
                stats['upcoming_demand'] += plant_data.get('demand', 0)

        plant_stats[plant.id] = {
            'plant': plant,
            'stats': stats,
        }

    # Build customers_by_plant: unique customer names per plant (for table iteration and display.)
    # OPTIMIZATION: Use pre-filtered IDs
    customers_by_plant = {}
    for plant in plants:
        names = []
        qs = simulation.customers.filter(plant=plant, id__in=filtered_customer_ids).order_by('name')
        for c in qs:
            if c.name not in names:
                names.append(c.name)
        customers_by_plant[plant.id] = names

    # Build refineries_by_plant: unique refinery names per plant
    # OPTIMIZATION: Use pre-filtered IDs
    refineries_by_plant = {}
    for plant in plants:
        names = []
        qs = simulation.refineries.filter(plant=plant, id__in=filtered_refinery_ids).order_by('name')
        for r in qs:
            if r.name not in names:
                names.append(r.name)
        refineries_by_plant[plant.id] = names

    # Build customers_summary: aggregate customers by plant and name
    # OPTIMIZATION: Use pre-filtered querysets
    customers_summary = {}
    for plant in plants:
        summary = {}
        qs = simulation.customers.filter(plant=plant, id__in=filtered_customer_ids).order_by('name')
        
        for c in qs.prefetch_related('date_ranges__supplier__plant'):
            range_list = []
            
            # Filter date ranges at Python level (smaller dataset now)
            for cd in c.date_ranges.all():
                if filter_from_date and filter_to_date:
                    if not (cd.from_date <= filter_to_date and cd.to_date >= filter_from_date):
                        continue
                
                supplier_data = None
                if cd.supplier:
                    supplier_data = {
                        'id': cd.supplier.id,
                        'name': cd.supplier.name,
                        'plant': cd.supplier.plant,
                    }
                
                range_dict = {
                    'id': cd.id,
                    'from_date': cd.from_date,
                    'to_date': cd.to_date,
                    'daily_demand': float(cd.daily_demand),
                    'supplier': supplier_data,
                }
                range_list.append(range_dict)
            
            if c.name in summary:
                summary[c.name]['ranges'].extend(range_list)
            else:
                summary[c.name] = {
                    'id': c.id,
                    'name': c.name,
                    'plant': plant,
                    'preference': c.preference,
                    'ranges': range_list
                }
        customers_summary[plant.id] = list(summary.values())
        
        # DEBUG: Log customer data for this plant
        if customers_summary[plant.id]:
            print(f"\nDEBUG - Customers for Plant {plant.name}:")
            for cust in customers_summary[plant.id]:
                print(f"  Customer: {cust['name']}")
                for r in cust['ranges']:
                    supplier_name = r['supplier']['name'] if r['supplier'] else 'NO SUPPLIER'
                    plant_name = r['supplier']['plant'].name if r['supplier'] and r['supplier']['plant'] else 'NO PLANT'
                    print(f"    - Demand: {r['daily_demand']} | Supplier: {supplier_name} ({plant_name})")
                    supplier_name = r['supplier']['name'] if r['supplier'] else 'NO SUPPLIER'
                    plant_name = r['supplier']['plant'].name if r['supplier'] and r['supplier']['plant'] else 'NO PLANT'
                    print(f"    - Demand: {r['daily_demand']} | Supplier: {supplier_name} ({plant_name})")

    # Build refineries_summary: aggregate refineries by plant and name
    refineries_summary = {}
    for plant in plants:
        summary = {}
        # Use pre-filtered IDs and select_related to fetch supplier data efficiently
        qs = simulation.refineries.filter(plant=plant, id__in=filtered_refinery_ids).order_by('name')
        
        # DEBUG: Log refinery query results
        print(f"\nDEBUG - Refineries for Plant {plant.name}:")
        print(f"  QuerySet count: {qs.count()}")
        
        for r in qs:
            print(f"  Processing refinery: {r.name} (ID: {r.id})")
            range_list = []
            # Prefetch supplier data with select_related - also prefetch the supplier's plant
            for rd in r.date_ranges.all().select_related('supplier__plant'):
                # Filter date ranges at Python level (smaller dataset now)
                if filter_from_date and filter_to_date:
                    if not (rd.from_date <= filter_to_date and rd.to_date >= filter_from_date):
                        continue
                
                # Create supplier dict to ensure template can access properties
                supplier_data = None
                if rd.supplier:
                    supplier_data = {
                        'id': rd.supplier.id,
                        'name': rd.supplier.name,
                        'plant': rd.supplier.plant,  # Keep as model for plant.name access
                    }
                
                range_dict = {
                    'id': rd.id,
                    'from_date': rd.from_date,
                    'to_date': rd.to_date,
                    'daily_refinery_demand': float(rd.daily_refinery_demand),
                    'supplier': supplier_data,  # Use dict instead of model instance
                }
                range_list.append(range_dict)
            
            if r.name in summary:
                summary[r.name]['ranges'].extend(range_list)
            else:
                print(f"    Adding refinery to summary: {r.name}")
                summary[r.name] = {
                    'id': r.id,
                    'name': r.name,
                    'plant': plant,
                    'preference': r.preference,
                    'ranges': range_list
                }
        refineries_summary[plant.id] = list(summary.values())
        print(f"  Final refineries_summary for Plant {plant.name}: {[r['name'] for r in refineries_summary[plant.id]]}")

    # Build suppliers_summary: aggregate suppliers by plant with date ranges
    # OPTIMIZATION: Use pre-filtered querysets
    suppliers_summary = {}
    for plant in plants:
        summary = {}
        qs = simulation.suppliers.filter(plant=plant, id__in=filtered_supplier_ids).order_by('name')
        
        for s in qs.prefetch_related('date_ranges'):
            range_list = []
            
            # Filter date ranges at Python level (smaller dataset now)
            for sd in s.date_ranges.all():
                if filter_from_date and filter_to_date:
                    if not (sd.from_date <= filter_to_date and sd.to_date >= filter_from_date):
                        continue
                
                range_dict = {
                    'id': sd.id,
                    'from_date': sd.from_date,
                    'to_date': sd.to_date,
                    'daily_supply': float(sd.daily_supply),
                }
                range_list.append(range_dict)
            
            if s.name in summary:
                summary[s.name]['ranges'].extend(range_list)
            else:
                summary[s.name] = {
                    'id': s.id,
                    'name': s.name,
                    'plant': plant,
                    'preference': s.preference,
                    'ranges': range_list
                }
        suppliers_summary[plant.id] = list(summary.values())

    # Get simulation comments ordered by newest first
    comments = simulation.comments.select_related('created_by').order_by('-created_at')

    # Cargos are already filtered via filtered_cargo_ids, no need for additional filtering
    
    return {
        # querysets used in supplier/cargo/customer tables - FILTERED!
        'suppliers': simulation.suppliers.filter(id__in=filtered_supplier_ids),
        'cargos': simulation.cargos.filter(id__in=filtered_cargo_ids),
        'customers': simulation.customers.filter(id__in=filtered_customer_ids),
        'refineries': simulation.refineries.filter(id__in=filtered_refinery_ids),
        'customers_by_plant': customers_by_plant,
        'refineries_by_plant': refineries_by_plant,
        'customers_summary': customers_summary,
        'refineries_summary': refineries_summary,
        'suppliers_summary': suppliers_summary,
        'plants':    plants,
        'daily_data':  daily_data,
        'plant_alerts': plant_alerts,
        'total_days':   len(daily_data),
        'comments': comments,

        # ── flat keys consumed by the "Total" stats panel in the template ──
        'total_opening_inventory':   total_opening_inventory,
        'total_supplied_till_today': total_supplied_till_today,
        'total_demand_till_today':   total_demand_till_today,
        'total_current_inventory':   total_current_inventory,
        'total_upcoming_supply':     total_upcoming_supply,
        'total_upcoming_demand':     total_upcoming_demand,

        # ── plant-wise dict consumed by the per-plant stat panels ──
        'plant_stats': plant_stats,
    }


def calculate_daily_data(simulation):
    # Build the timeline of inventory, supply, and demand for each plant.
    # The projection iterates one day at a time from simulation start to end.
    # 
    # OPTIMIZED VERSION: Pre-fetches all data to avoid N+1 query problem
    # 
    # UPDATED BUSINESS FLOW (Two Separate Inventories):
    # 1. Supplier Supply → Customer Demand (customers served directly from supplier)
    # 2. Remaining Supplier = Total Supplier - Total Customer Demand
    # 3. If Supplier < Customer Demand: Shortfall served from Cargo Inventory
    # 4. Plant Inventory = Opening + Remaining Supplier (NEVER decreases from demand)
    # 5. Cargo Inventory = Previous + Cargo Arrivals - Shortfall (can go negative)
    # 6. Closing Inventory = Plant Inventory + Cargo Inventory
    
    from django.db.models import Prefetch
    
    start_date = simulation.start_date
    end_date = simulation.end_date
    num_days = (end_date - start_date).days + 1
    
    # ── OPTIMIZATION: Pre-fetch ALL related data in single queries ────────────
    # Get all plants with their inventories
    plants_qs = simulation.plant_inventories.select_related('plant').all()
    plant_inventories = {}
    cargo_inventories = {}
    
    for pi in plants_qs:
        plant_inventories[pi.plant_id] = float(pi.opening_inventory)
        cargo_inventories[pi.plant_id] = 0.0  # Start with 0 cargo inventory
    
    # Pre-fetch all suppliers with their date ranges (single query with JOIN)
    suppliers_qs = simulation.suppliers.filter(plant_id__in=plant_inventories.keys()).select_related('plant').prefetch_related(
        Prefetch('date_ranges', queryset=SupplierDate.objects.all().order_by('from_date'))
    )
    
    # Pre-fetch all customers with their date ranges and suppliers
    customers_qs = simulation.customers.filter(plant_id__in=plant_inventories.keys()).select_related('plant').prefetch_related(
        Prefetch('date_ranges', queryset=CustomerDate.objects.select_related('supplier').all().order_by('from_date'))
    )
    
    # Pre-fetch all refineries with their date ranges and suppliers
    refineries_qs = simulation.refineries.filter(plant_id__in=plant_inventories.keys()).select_related('plant').prefetch_related(
        Prefetch('date_ranges', queryset=RefineryDate.objects.select_related('supplier').all().order_by('from_date'))
    )
    
    # Pre-fetch all cargos for the simulation period
    cargos_qs = simulation.cargos.filter(delivery_date__gte=start_date, delivery_date__lte=end_date).all()
    
    # Build lookup dictionaries for O(1) access instead of database queries
    suppliers_by_plant = {}
    for supplier in suppliers_qs:
        if supplier.plant_id not in suppliers_by_plant:
            suppliers_by_plant[supplier.plant_id] = []
        suppliers_by_plant[supplier.plant_id].append(supplier)
    
    customers_by_plant = {}
    for customer in customers_qs:
        if customer.plant_id not in customers_by_plant:
            customers_by_plant[customer.plant_id] = []
        customers_by_plant[customer.plant_id].append(customer)
    
    refineries_by_plant = {}
    for refinery in refineries_qs:
        if refinery.plant_id not in refineries_by_plant:
            refineries_by_plant[refinery.plant_id] = []
        refineries_by_plant[refinery.plant_id].append(refinery)
    
    cargos_by_date = {}
    for cargo in cargos_qs:
        if cargo.delivery_date not in cargos_by_date:
            cargos_by_date[cargo.delivery_date] = 0.0
        cargos_by_date[cargo.delivery_date] += float(cargo.amount)
    
    daily_data = []
    current_date = start_date
    
    # Iterate through every simulation day, creating a summary record for each date.
    while current_date <= end_date:
        day_data = {
            'date': current_date,
            'plant_data': {},
            'total_supply': 0,
            'total_customer_demand': 0,
            'total_remaining_supply': 0,
            'total_cargo_arrival': 0,
            'total_inventory_available': 0,
            'total_refinery_demand': 0,
            'total_demand': 0,
            'total_requested_demand': 0,
            'total_plant_inventory': 0,
            'total_cargo_inventory': 0,
            'total_inventory': 0,
            'total_projected_balance': 0
        }
        
        for plant_id, opening_inventory in plant_inventories.items():
            # Reset per-plant daily totals
            total_supplier_supply = 0.0
            total_customer_demand = 0.0
            total_cargo = 0.0
            total_refinery_demand = 0.0
            
            supplies = []
            demands = []
            
            # Aggregated entity values for this day (per entity name)
            suppliers_today = {}  # {supplier_name: total_daily_supply}
            customers_today = {}  # {customer_name: {'total_demand': X, 'preference': Y, 'supplier': Z}}
            refineries_today = {}  # {refinery_name: total_daily_demand}
            
            # Supplier allocation breakdown for this day
            supplier_allocation = {}  # {supplier_name: {'supply': X, 'demands': [{customer: Y}], 'remaining': Z}}
            
            # ── OPTIMIZED STEP 1: Calculate Total Supplier Supply (using pre-fetched data) ────────────────
            for supplier in suppliers_by_plant.get(plant_id, []):
                supplier_daily_supply = 0.0
                for sd in supplier.date_ranges.all():
                    if sd.from_date <= current_date <= sd.to_date:
                        amount = float(sd.daily_supply)
                        supplies.append({
                            'type': 'supplier', 
                            'name': supplier.name, 
                            'amount': amount,
                            'range_id': sd.id
                        })
                        supplier_daily_supply += amount
                
                if supplier_daily_supply > 0:
                    total_supplier_supply += supplier_daily_supply
                    # Aggregate by supplier name
                    if supplier.name not in suppliers_today:
                        suppliers_today[supplier.name] = 0.0
                        # Initialize supplier allocation tracking
                        supplier_allocation[supplier.name] = {
                            'supply': 0.0,
                            'demands': [],
                            'remaining': 0.0
                        }
                    suppliers_today[supplier.name] += supplier_daily_supply
                    supplier_allocation[supplier.name]['supply'] += supplier_daily_supply
            
            # ── STEP 2: Serve Connected Customers from Their Assigned Suppliers ────────────────
            # BUSINESS RULE: Each customer is served FIRST from its assigned supplier
            total_customer_demand = 0.0
            customer_shortfall_total = 0.0
            
            for customer in customers_by_plant.get(plant_id, []):
                customer_daily_demand = 0.0
                assigned_supplier_name = None
                
                # Get the active date range and its assigned supplier
                for cd in customer.date_ranges.all():
                    if cd.from_date <= current_date <= cd.to_date:
                        customer_daily_demand += float(cd.daily_demand)
                        # Track supplier assignment from CustomerDate
                        if cd.supplier:
                            assigned_supplier_name = cd.supplier.name
                
                if customer_daily_demand > 0:
                    total_customer_demand += customer_daily_demand
                    
                    # Aggregate by customer name
                    if customer.name not in customers_today:
                        customers_today[customer.name] = {
                            'total_demand': 0.0,
                            'preference': customer.preference,
                            'supplier': assigned_supplier_name
                        }
                    customers_today[customer.name]['total_demand'] += customer_daily_demand
                    customers_today[customer.name]['preference'] = min(
                        customers_today[customer.name]['preference'],
                        customer.preference
                    )
                    
                    # Track demand in supplier allocation if supplier is assigned
                    if assigned_supplier_name and assigned_supplier_name in supplier_allocation:
                        supplier_allocation[assigned_supplier_name]['demands'].append({
                            'customer': customer.name,
                            'demand': customer_daily_demand,
                            'type': 'customer'
                        })
            
            # ── STEP 3: Serve Connected Refineries from Their Assigned Suppliers ────────────────
            # BUSINESS RULE: Each refinery is served FIRST from its assigned supplier
            total_refinery_demand = 0.0
            refinery_shortfall_total = 0.0
            
            print(f"\n🔍 Processing refineries for plant {plant_id} on {current_date}")
            
            for refinery in refineries_by_plant.get(plant_id, []):
                refinery_daily_demand = 0.0
                assigned_supplier_name = None
                
                # Get the active date range and its assigned supplier
                for rd in refinery.date_ranges.all():
                    if rd.from_date <= current_date <= rd.to_date:
                        refinery_daily_demand += float(rd.daily_refinery_demand)
                        # Track supplier assignment from RefineryDate
                        if rd.supplier:
                            assigned_supplier_name = rd.supplier.name
                
                print(f"  Refinery: {refinery.name}, Demand: {refinery_daily_demand}, Supplier: {assigned_supplier_name}")
                
                if refinery_daily_demand > 0:
                    total_refinery_demand += refinery_daily_demand
                    
                    # Aggregate by refinery name
                    if refinery.name not in refineries_today:
                        refineries_today[refinery.name] = {
                            'total_demand': 0.0,
                            'preference': refinery.preference,
                            'supplier': assigned_supplier_name
                        }
                    refineries_today[refinery.name]['total_demand'] += refinery_daily_demand
                    refineries_today[refinery.name]['preference'] = min(
                        refineries_today[refinery.name]['preference'],
                        refinery.preference
                    )
                    
                    # Track demand in supplier allocation if supplier is assigned
                    if assigned_supplier_name and assigned_supplier_name in supplier_allocation:
                        supplier_allocation[assigned_supplier_name]['demands'].append({
                            'refinery': refinery.name,
                            'demand': refinery_daily_demand,
                            'type': 'refinery'
                        })
            
            print(f"  ✅ Total Refinery Demand for plant {plant_id}: {total_refinery_demand}")
            print(f"  ✅ Refineries today: {list(refineries_today.keys())}")
            
            # ── STEP 4: Calculate Supplier Remaining After All Allocations ────────────────
            # For each supplier: Remaining = Supply - (Customer Demand + Refinery Demand served)
            for supplier_name, alloc in supplier_allocation.items():
                total_demand_from_this_supplier = sum(d['demand'] for d in alloc['demands'])
                alloc['remaining'] = max(0.0, alloc['supply'] - total_demand_from_this_supplier)
            
            # ── STEP 5: Calculate Shortfalls (Demand NOT served by supplier) ────────────────
            # Customer shortfall = Customer demand that couldn't be served by assigned suppliers OR cargo
            # Refinery shortfall = Refinery demand that couldn't be served by assigned suppliers OR cargo
            
            customer_shortfall_total = 0.0
            refinery_shortfall_total = 0.0
            
            # Track cumulative cargo withdrawals for this day (to ensure we don't over-withdraw)
            total_cargo_needed_for_unassigned = 0.0
            
            for customer_name, data in customers_today.items():
                supplier_name = data.get('supplier')
                if supplier_name and supplier_name in supplier_allocation:
                    # Calculate how much of this customer's demand was served by supplier
                    total_customer_demand_for_entity = data['total_demand']
                    
                    # Find all demands from this supplier for this customer
                    demands_from_supplier = [d for d in supplier_allocation[supplier_name]['demands'] 
                                            if d.get('customer') == customer_name]
                    demand_to_allocate = sum(d['demand'] for d in demands_from_supplier)
                    
                    # Customer served = min(customer_demand, supplier_remaining_before_this_customer)
                    # For simplicity: serve proportionally from available supplier supply
                    supplier_available = supplier_allocation[supplier_name]['supply'] - \
                                        sum(d['demand'] for d in supplier_allocation[supplier_name]['demands'] 
                                           if d.get('type') == 'refinery')
                    
                    served_to_customer = min(total_customer_demand_for_entity, supplier_available)
                    remaining_after_supplier = max(0.0, total_customer_demand_for_entity - served_to_customer)
                    
                    # If supplier didn't cover full demand, try to use cargo inventory
                    if remaining_after_supplier > 0:
                        total_cargo_needed_for_unassigned += remaining_after_supplier
                else:
                    # No supplier assigned - demand will be served from cargo inventory
                    total_cargo_needed_for_unassigned += data['total_demand']
            
            for refinery_name, data in refineries_today.items():
                supplier_name = data.get('supplier')
                if supplier_name and supplier_name in supplier_allocation:
                    total_refinery_demand_for_entity = data['total_demand']
                    
                    # Find all demands from this supplier for this refinery
                    demands_from_supplier = [d for d in supplier_allocation[supplier_name]['demands'] 
                                            if d.get('refinery') == refinery_name]
                    demand_to_allocate = sum(d['demand'] for d in demands_from_supplier)
                    
                    # Calculate remaining supplier after customer allocations
                    supplier_remaining_after_customers = supplier_allocation[supplier_name]['supply'] - \
                        sum(d['demand'] for d in supplier_allocation[supplier_name]['demands'] 
                           if d.get('type') == 'customer')
                    
                    served_to_refinery = min(total_refinery_demand_for_entity, max(0.0, supplier_remaining_after_customers))
                    remaining_after_supplier = max(0.0, total_refinery_demand_for_entity - served_to_refinery)
                    
                    # If supplier didn't cover full demand, try to use cargo inventory
                    if remaining_after_supplier > 0:
                        total_cargo_needed_for_unassigned += remaining_after_supplier
                else:
                    # No supplier assigned - demand will be served from cargo inventory
                    total_cargo_needed_for_unassigned += data['total_demand']
            
            # Now calculate actual shortfall based on available cargo inventory
            # We'll determine this after calculating cargo availability in Step 8
            # For now, mark all unassigned/partially-served demand as needing cargo
            
            # ── STEP 6: Add Remaining Supplier to Plant Inventory ────────────────
            # BUSINESS RULE: Plant Inventory can go negative (no longer capped at 0)
            remaining_supplier_after_all_demands = sum(alloc['remaining'] for alloc in supplier_allocation.values())
            plant_inventories[plant_id] += remaining_supplier_after_all_demands
            
            # ── STEP 7: Add Cargo Arrivals BEFORE serving demand ────────────────
            total_cargo_arrival = cargos_by_date.get(current_date, 0.0)
            
            # Track cargo inventory at each step for display
            cargo_before_arrival = cargo_inventories[plant_id]
            
            if total_cargo_arrival > 0:
                supplies.append({'type': 'cargo', 'name': f"Cargo (Total)", 'amount': total_cargo_arrival})
                cargo_inventories[plant_id] += total_cargo_arrival
            
            cargo_after_arrival = cargo_inventories[plant_id]
            
            # ── STEP 8: Serve Unassigned/Partially-Served Demand from Cargo Inventory ────────────────
            # BUSINESS RULE: If no supplier is linked (or supplier didn't cover full demand),
            # automatically take from cargo inventory first. Cargo can go negative!
            
            # Calculate how much cargo we actually have available before this withdrawal
            available_cargo_before_withdrawal = cargo_inventories[plant_id]
            
            # Withdraw FULL amount needed from cargo (can make it very negative!)
            actual_cargo_withdrawal = total_cargo_needed_for_unassigned
            cargo_inventories[plant_id] -= actual_cargo_withdrawal
            
            # Track final cargo after withdrawal (will be negative if withdrawal > available)
            cargo_after_withdrawal = cargo_inventories[plant_id]
            
            # Shortfall is now 0 because we allowed cargo to go negative to cover all demand
            total_shortfall = 0.0
            
            # Distribute shortfall proportionally between customers and refineries based on their unmet demand
            if total_cargo_needed_for_unassigned > 0:
                customer_ratio = sum(data['total_demand'] for data in customers_today.values() 
                                    if not data.get('supplier') or data.get('supplier') not in supplier_allocation) / total_cargo_needed_for_unassigned
                refinery_ratio = sum(data['total_demand'] for data in refineries_today.values() 
                                    if not data.get('supplier') or data.get('supplier') not in supplier_allocation) / total_cargo_needed_for_unassigned
            else:
                customer_ratio = 0.5
                refinery_ratio = 0.5
            
            # Since cargo can go negative, there's no shortfall - all demand is fully served!
            customer_shortfall_total = 0.0
            refinery_shortfall_total = 0.0
            
            # ── STEP 9: Calculate Closing Inventory ────────────────
            # BUSINESS RULE: Closing Inventory = Plant Inventory + Cargo Inventory
            closing_inventory = plant_inventories[plant_id] + cargo_inventories[plant_id]
            
            # Projected Inventory Balance (uncapped, can be negative)
            projected_balance = closing_inventory
            
            # ── STEP 10: Build Demands List for Display ────────────────
            # Since cargo can go negative, ALL demand is fully served (no shortfall)
            
            for customer_name, data in customers_today.items():
                demands.append({
                    'source': customer_name,
                    'served_amount': data['total_demand'],  # Full demand served (from supplier + possibly negative cargo)
                    'requested_amount': data['total_demand'],
                    'preference': data['preference'],
                    'type': 'customer',
                    'plant': plant_id,
                    'supplier': data.get('supplier')
                })
            
            for refinery_name, data in refineries_today.items():
                demands.append({
                    'source': refinery_name,
                    'served_amount': data['total_demand'],  # Full demand served (from supplier + possibly negative cargo)
                    'requested_amount': data['total_demand'],
                    'preference': data['preference'],
                    'type': 'refinery',
                    'plant': plant_id,
                    'supplier': data.get('supplier')
                })
            
            # No unserved demand since we allow negative cargo
            unserved_demand = 0.0
            backlog = 0.0  # No backlog since inventory can go negative
            
            # Store the closing inventory for this plant after serving demand
            day_data['plant_data'][plant_id] = {
                'supplies': supplies,
                'demands': demands,
                'total_supplier_supply': total_supplier_supply,
                'total_customer_demand': total_customer_demand,
                'total_refinery_demand': total_refinery_demand,  # ADDED: Per-plant refinery demand for display
                'customer_shortfall': customer_shortfall_total,
                'refinery_shortfall': refinery_shortfall_total,
                'total_cargo_withdrawal': actual_cargo_withdrawal,
                'total_cargo_arrival': total_cargo_arrival,
                'cargo_before_arrival': cargo_before_arrival,  # For display/debugging
                'cargo_after_arrival': cargo_after_arrival,     # For display/debugging
                'cargo_after_withdrawal': cargo_after_withdrawal, # For display/debugging
                'plant_inventory': plant_inventories[plant_id],
                'cargo_inventory': cargo_inventories[plant_id],
                'supply': remaining_supplier_after_all_demands + total_cargo_arrival,
                'demand': total_customer_demand + total_refinery_demand,
                'requested_demand': total_customer_demand + total_refinery_demand,
                'unserved_demand': unserved_demand,
                'inventory': closing_inventory,
                'closing_inventory': closing_inventory,  # No longer capped at 0
                'backlog': backlog,
                'is_negative': closing_inventory < 0,
                'suppliers_today': suppliers_today,
                'customers_today': customers_today,
                'refineries_today': refineries_today,
                'supplier_allocation': supplier_allocation,
                'projected_balance': projected_balance
            }
            
            day_data['total_supply'] += total_supplier_supply
            day_data['total_customer_demand'] += total_customer_demand
            # Calculate remaining supplier after ALL demands (customer + refinery)
            remaining_after_all = sum(alloc['remaining'] for alloc in supplier_allocation.values())
            day_data['total_remaining_supply'] += remaining_after_all
            day_data['total_cargo_arrival'] += total_cargo_arrival
            day_data['total_refinery_demand'] += total_refinery_demand
            # Total demand is actual requested demand
            day_data['total_demand'] += total_customer_demand + total_refinery_demand
            day_data['total_requested_demand'] += total_customer_demand + total_refinery_demand
            day_data['total_plant_inventory'] += plant_inventories[plant_id]
            day_data['total_cargo_inventory'] += cargo_inventories[plant_id]
            day_data['total_inventory'] += closing_inventory
        
        day_data['total_backlog'] = max(0.0, -day_data['total_inventory'])
        # Add total projected balance for all plants (uncapped)
        day_data['total_projected_balance'] = sum(
            day_data['plant_data'][pid].get('projected_balance', 0) 
            for pid in day_data['plant_data']
        )
        daily_data.append(day_data)
        current_date += timedelta(days=1)
    
    return daily_data


def initialize_sample_data(simulation):
    """Initialize simulation with sample data"""
    plants = Plant.objects.all()
    if not plants.exists():
        return
    
    plant_list = list(plants[:2])
    if len(plant_list) < 2:
        plant_list = plant_list * 2
    
    # Create suppliers with date ranges
    supplier_a = Supplier.objects.create(
        simulation=simulation, plant=plant_list[0], name='Supplier A', preference=1
    )
    SupplierDate.objects.create(
        supplier=supplier_a,
        from_date=simulation.start_date, 
        to_date=simulation.end_date,
        daily_supply=100
    )
    
    supplier_b = Supplier.objects.create(
        simulation=simulation, plant=plant_list[1], name='Supplier B', preference=1
    )
    SupplierDate.objects.create(
        supplier=supplier_b,
        from_date=simulation.start_date, 
        to_date=simulation.start_date + timedelta(days=180),
        daily_supply=120
    )
    
    Cargo.objects.create(
        simulation=simulation, plant=plant_list[0], cargo_name='Cargo-001',
        delivery_date=simulation.start_date + timedelta(days=15), amount=5000
    )
    Cargo.objects.create(
        simulation=simulation, plant=plant_list[1], cargo_name='Cargo-002',
        delivery_date=simulation.start_date + timedelta(days=80), amount=4500
    )
    
    # Create customers with date ranges
    customer_x = Supplier.objects.create(
        simulation=simulation, plant=plant_list[0], name='Customer X', preference=1
    )
    CustomerDate.objects.create(
        customer=customer_x,
        from_date=simulation.start_date, 
        to_date=simulation.end_date,
        daily_demand=80
    )
    
    customer_y = Customer.objects.create(
        simulation=simulation, plant=plant_list[1], name='Customer Y', preference=1
    )
    CustomerDate.objects.create(
        customer=customer_y,
        from_date=simulation.start_date, 
        to_date=simulation.start_date + timedelta(days=180),
        daily_demand=70
    )


# ── Master simulation administration ───────────────────────────────────────

@login_required
def manage_master_simulation(request):
    # Administrator view for creating or editing the central master simulation.
    # The master simulation serves as the template for new user simulations.
    if not is_planning_admin(request.user):
        messages.error(request, 'Only administrators can manage master simulation')
        return redirect('lng_planner:dashboard')
    
    master = Simulation.objects.filter(is_master=True).first()
    
    if request.method == 'POST':
        form = MasterSimulationForm(request.POST, instance=master)
        if form.is_valid():
            master_sim = form.save(commit=False)
            master_sim.user = None
            master_sim.is_master = True
            master_sim.is_active = False
            master_sim.save()
            
            messages.success(request, 'Master simulation saved successfully!')
            return redirect('lng_planner:setup_plants', simulation_id=master_sim.id)
    else:
        form = MasterSimulationForm(instance=master)
    
    return render(request, 'lng_planner/manage_master.html', {
        'form': form,
        'master': master,
        'title': 'Edit Master Simulation' if master else 'Create Master Simulation'
    })


@login_required
def copy_from_master(request):
    # Create a personal simulation by duplicating master simulation data.
    print("\n" + "="*80)
    print("🚨🚨🚨 COPY_FROM_MASTER FUNCTION CALLED 🚨🚨🚨")
    print(f"  User: {request.user.username}")
    print(f"  Time: {datetime.now()}")
    print("="*80 + "\n")
    
    master = Simulation.objects.filter(is_master=True).first()
    
    if not master:
        messages.error(request, 'No master simulation found!')
        return redirect('lng_planner:dashboard')
    
    # GET request - show form for custom name
    if request.method == 'GET':
        suggested_name = f"{master.name} - Copy {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        context = {
            'master': master,
            'suggested_name': suggested_name,
        }
        return render(request, 'lng_planner/create_from_master.html', context)
    
    # POST request - create simulation with custom name from form
    sim_name = request.POST.get('name', f"{master.name} - Copy {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    
    # First, deactivate all existing user simulations
    Simulation.objects.filter(user=request.user).update(is_active=False)
    
    # Create new simulation with is_active=True
    new_sim = Simulation.objects.create(
        user=request.user,
        name=sim_name,
        start_date=master.start_date,
        end_date=master.end_date,
        is_master=False,
        is_active=True  # This will be the active simulation
    )
    
    print(f"\n✅ Created new simulation: {new_sim.name} (ID: {new_sim.id}, is_active: {new_sim.is_active})")
    
    for pi in master.plant_inventories.all():
        PlantInventory.objects.create(
            simulation=new_sim, plant=pi.plant, opening_inventory=pi.opening_inventory
        )
    
    Simulation.objects.filter(user=request.user).exclude(pk=new_sim.pk).update(is_active=False)
    
    # Copy suppliers with date ranges
    print(f"\n📦 COPYING SUPPLIERS:")
    for supplier in master.suppliers.all():
        new_supplier = Supplier.objects.create(
            simulation=new_sim, plant=supplier.plant, name=supplier.name,
            preference=supplier.preference
        )
        # Copy date ranges
        for sd in supplier.date_ranges.all():
            SupplierDate.objects.create(
                supplier=new_supplier,
                from_date=sd.from_date, to_date=sd.to_date,
                daily_supply=sd.daily_supply
            )
        print(f"  ✅ Copied supplier: {supplier.name} with {supplier.date_ranges.count()} date ranges")
    
    for cargo in master.cargos.all():
        Cargo.objects.create(
            simulation=new_sim, plant=cargo.plant, cargo_name=cargo.cargo_name,
            delivery_date=cargo.delivery_date, amount=cargo.amount
        )
    
    # Copy customers with date ranges (including supplier relationships)
    print(f"\n{'='*60}")
    print(f"📋 COPYING FROM MASTER - Customer Data:")
    print(f"  Master Simulation ID: {master.id}")
    print(f"  Total Customers in Master: {master.customers.count()}")
    
    # DIRECT DATABASE CHECK - Query CustomerDate table directly
    from lng_planner.models import CustomerDate as CD
    master_customer_dates = CD.objects.filter(customer__simulation=master)
    print(f"\n  🔍 DIRECT DB QUERY: Total CustomerDate records in Master: {master_customer_dates.count()}")
    for cd in master_customer_dates[:5]:  # Show first 5
        supplier_name = cd.supplier.name if cd.supplier else "NULL"
        print(f"    CD ID:{cd.id} | Customer:{cd.customer.name} | Supplier: {supplier_name}")
    
    for customer in master.customers.all():
        customer_date_count = customer.date_ranges.count()
        print(f"\n  Customer: {customer.name} (Plant: {customer.plant.name})")
        print(f"    CustomerDate records in Master: {customer_date_count}")
        
        # List all CustomerDate records for this customer in master
        for cd in customer.date_ranges.all():
            supplier_info = f"Supplier: {cd.supplier.name} ({cd.supplier.plant.name})" if cd.supplier else "NO SUPPLIER"
            print(f"      - ID:{cd.id} | {cd.from_date} to {cd.to_date}, Demand: {cd.daily_demand} | {supplier_info}")
        
        new_customer = Customer.objects.create(
            simulation=new_sim, plant=customer.plant, name=customer.name,
            preference=customer.preference
        )
        
        # Copy each CustomerDate record
        copied_count = 0
        for cd in customer.date_ranges.all():
            supplier_info = f"Supplier: {cd.supplier.name} ({cd.supplier.plant.name})" if cd.supplier else "NO SUPPLIER"
            print(f"    → Copying: {cd.from_date} to {cd.to_date}, Demand: {cd.daily_demand} | {supplier_info}")
            CustomerDate.objects.create(
                customer=new_customer,
                supplier=cd.supplier,  # Preserve the supplier relationship from master
                from_date=cd.from_date, 
                to_date=cd.to_date,
                daily_demand=cd.daily_demand
            )
            copied_count += 1
        
        print(f"    ✅ Copied {copied_count} CustomerDate records to new simulation")
    print(f"{'='*60}\n")
    
    # Copy refineries with date ranges
    for refinery in master.refineries.all():
        new_refinery = Refinery.objects.create(
            simulation=new_sim, plant=refinery.plant, name=refinery.name,
            preference=refinery.preference
        )
        for rd in refinery.date_ranges.all():
            RefineryDate.objects.create(
                refinery=new_refinery,
                from_date=rd.from_date, to_date=rd.to_date,
                daily_refinery_demand=rd.daily_refinery_demand
            )
    
    messages.success(request, f'Simulation "{sim_name}" created successfully from master!')
    return redirect('lng_planner:dashboard')


@login_required
def refresh_master_from_sap(request):
    """
    Admin-only process that refreshes master simulation from SAP API (or mock data).
    Creates a new MasterVersion and new Master Simulation, preserving all historical data.
    Uses mock data by default for testing; falls back to real SAP API if configured and available.
    """
    if not is_planning_admin(request.user):
        messages.error(request, 'Only administrators can refresh master data')
        return redirect('lng_planner:dashboard')
    
    # Get current active master simulation
    current_master = Simulation.objects.filter(is_master=True).first()
    
    if not current_master:
        messages.error(request, 'No master simulation found. Please create one first.')
        return redirect('lng_planner:manage_master')
    
    # Try to fetch from SAP API, but use mock data if unavailable
    sap_data = None
    api_used = False
    
    if current_master.sap_api_url:
        try:
            print(f"\n{'='*60}")
            print(f"🔄 REFRESH MASTER FROM SAP - Starting")
            print(f"  User: {request.user.username} ({'Planning Admin'})")
            print(f"  Current Master: {current_master.name} (ID: {current_master.id})")
            print(f"  SAP API URL: {current_master.sap_api_url}")
            
            # Try to fetch from real SAP API
            response = requests.get(current_master.sap_api_url, timeout=30)
            response.raise_for_status()
            sap_data = response.json()
            api_used = True
            
            print(f"  ✅ Successfully fetched data from SAP API")
            print(f"{'='*60}\n")
            
        except Exception as e:
            print(f"\n⚠️  SAP API failed ({type(e).__name__}: {str(e)})")
            print(f"  → Falling back to mock data for testing\n")
            messages.warning(request, f'⚠️ Could not connect to SAP API - using mock data: {str(e)}')
    
    # If no SAP data or URL configured, use mock data
    if not sap_data:
        sap_data = get_mock_sap_data()
        api_used = False
        print(f"\n{'='*60}")
        print(f"🔄 REFRESH MASTER FROM MOCK DATA")
        print(f"  User: {request.user.username} ({'Planning Admin'})")
        print(f"  Current Master: {current_master.name} (ID: {current_master.id})")
        print(f"  Using mock data for testing")
        print(f"{'='*60}\n")
    
    # Determine next version number - find highest numeric version
    all_sap_versions = MasterVersion.objects.filter(source_type='SAP').order_by('-created_at')
    next_num = 1
    
    for mv in all_sap_versions:
        try:
            num = int(mv.version_number.replace('V', ''))
            if num >= next_num:
                next_num = num + 1
            break  # Only check the most recent one
        except ValueError:
            continue  # Skip invalid version numbers
    
    version_number = f"V{next_num}"
    
    # Ensure this version number doesn't already exist (safety check)
    while MasterVersion.objects.filter(version_number=version_number).exists():
        next_num += 1
        version_number = f"V{next_num}"
    
    print(f"Creating new MasterVersion: {version_number}")
    
    # Create new MasterVersion
    new_master_version = MasterVersion.objects.create(
        version_number=version_number,
        name=f"SAP Sync - {datetime.now().strftime('%Y-%m-%d')}",
        source_type='SAP',
        created_by=request.user,
        description=f'Master data refreshed from {"SAP API" if api_used else "mock data"} on {datetime.now().strftime("%Y-%m-%d %H:%M")}',
        is_active=True  # This will be the active version
    )
    
    print(f"MasterVersion created: {new_master_version.id}")
    
    # Mark previous active MasterVersion as inactive
    MasterVersion.objects.filter(source_type='SAP', is_active=True).exclude(pk=new_master_version.pk).update(is_active=False)
    MasterVersion.objects.filter(source_type='SIMULATION', is_active=True).update(is_active=False)
    
    print(f"Old master versions deactivated")
    
    # Create NEW Master Simulation (preserve old one for history)
    new_master = Simulation.objects.create(
        user=None,
        name=f"Master - {version_number} ({datetime.now().strftime('%Y-%m-%d')})",
        start_date=current_master.start_date,
        end_date=current_master.end_date,
        is_master=True,
        is_active=True,
        sap_api_url=current_master.sap_api_url if api_used else '',
        master_version=new_master_version
    )
    
    print(f"New Master Simulation created: {new_master.name} (ID: {new_master.id})")
    
    # Mark old master as inactive (but keep it for history)
    current_master.is_active = False
    current_master.save()
    
    print(f"Old master deactivated: {current_master.name}")
    
    counts = {
        'plants_created': 0, 
        'plants_existing': 0,
        'plant_inventories': 0, 
        'suppliers': 0,
        'cargos': 0, 
        'customers': 0, 
        'refineries': 0, 
        'errors': []
    }
    
    def get_or_create_plant(plant_name, location=''):
        """Get or create plant record and track creation stats"""
        plant, created = Plant.objects.get_or_create(
            name=plant_name, defaults={'location': location}
        )
        if created:
            counts['plants_created'] += 1
        else:
            counts['plants_existing'] += 1
        return plant
    
    # Helper to parse date strings
    def parse_date(date_str):
        """Parse date string in YYYY-MM-DD format"""
        from datetime import datetime
        return datetime.strptime(date_str, '%Y-%m-%d').date()
    
    # Import Plant Inventories
    if 'plant_inventories' in sap_data:
        for pi_data in sap_data['plant_inventories']:
            try:
                plant = get_or_create_plant(pi_data['plant_name'], pi_data.get('location', ''))
                PlantInventory.objects.create(
                    simulation=new_master, 
                    plant=plant,
                    opening_inventory=float(pi_data['opening_inventory'])
                )
                counts['plant_inventories'] += 1
            except Exception as e:
                counts['errors'].append(f"Plant inventory error: {str(e)}")
    
    # Import Suppliers with date ranges
    if 'suppliers' in sap_data:
        for s in sap_data['suppliers']:
            try:
                plant = get_or_create_plant(s['plant_name'], s.get('location', ''))
                supplier = Supplier.objects.create(
                    simulation=new_master, 
                    plant=plant, 
                    name=s['name'],
                    preference=s.get('preference', 1)
                )
                
                # Handle date_ranges array or single from_date/to_date
                if 'date_ranges' in s:
                    for dr in s['date_ranges']:
                        SupplierDate.objects.create(
                            supplier=supplier,
                            from_date=parse_date(dr['from_date']), 
                            to_date=parse_date(dr['to_date']),
                            daily_supply=float(dr['daily_supply'])
                        )
                else:
                    # Legacy format with single date range
                    SupplierDate.objects.create(
                        supplier=supplier,
                        from_date=parse_date(s['from_date']), 
                        to_date=parse_date(s['to_date']),
                        daily_supply=float(s['daily_supply'])
                    )
                counts['suppliers'] += 1
            except Exception as e:
                counts['errors'].append(f"Supplier error ({s.get('name', 'unknown')}): {str(e)}")
    
    # Import Cargos
    if 'cargos' in sap_data:
        for c in sap_data['cargos']:
            try:
                plant = get_or_create_plant(c['plant_name'], c.get('location', ''))
                Cargo.objects.create(
                    simulation=new_master, 
                    plant=plant,
                    cargo_name=c['cargo_name'], 
                    delivery_date=parse_date(c['delivery_date']), 
                    amount=float(c['amount'])
                )
                counts['cargos'] += 1
            except Exception as e:
                counts['errors'].append(f"Cargo error ({c.get('cargo_name', 'unknown')}): {str(e)}")
    
    # Import Customers with date ranges and supplier mappings
    if 'customers' in sap_data:
        for c in sap_data['customers']:
            try:
                plant = get_or_create_plant(c['plant_name'], c.get('location', ''))
                customer = Customer.objects.create(
                    simulation=new_master, 
                    plant=plant, 
                    name=c['name'],
                    preference=c.get('preference', 1)
                )
                
                print(f"\n📝 Processing Customer: {c['name']} at {plant.name}")
                
                # Handle date_ranges array
                if 'date_ranges' in c:
                    for dr in c['date_ranges']:
                        # Find the supplier by name (need to get it from new_master)
                        supplier = None
                        supplier_name = dr.get('supplier_name', '')
                        
                        print(f"  Date Range: {dr['from_date']} to {dr['to_date']}, Demand: {dr['daily_demand']}")
                        
                        if supplier_name:
                            print(f"    🔍 Looking for Supplier: '{supplier_name}' at plant {plant.name}...")
                            supplier = Supplier.objects.filter(
                                simulation=new_master,
                                plant=plant,
                                name=supplier_name
                            ).first()
                            
                            if supplier:
                                print(f"    ✅ FOUND Supplier ID {supplier.id}: '{supplier_name}'")
                            else:
                                # DEBUG: List all suppliers at this plant
                                all_suppliers = list(Supplier.objects.filter(simulation=new_master, plant=plant).values_list('name', flat=True))
                                print(f"    ❌ NOT FOUND! Available suppliers at {plant.name}: {all_suppliers}")
                        else:
                            print(f"    ⚠️  No supplier_name in date range")
                        
                        CustomerDate.objects.create(
                            customer=customer,
                            supplier=supplier,
                            from_date=parse_date(dr['from_date']), 
                            to_date=parse_date(dr['to_date']),
                            daily_demand=float(dr['daily_demand'])
                        )
                    print(f"  ✅ Created {len(c['date_ranges'])} CustomerDate records for '{c['name']}'")
                else:
                    # Legacy format
                    supplier = None
                    if 'supplier_name' in c and c['supplier_name']:
                        supplier = Supplier.objects.filter(
                            simulation=new_master,
                            plant=plant,
                            name=c['supplier_name']
                        ).first()
                    
                    CustomerDate.objects.create(
                        customer=customer,
                        supplier=supplier,
                        from_date=parse_date(c['from_date']), 
                        to_date=parse_date(c['to_date']),
                        daily_demand=float(c['daily_demand'])
                    )
                counts['customers'] += 1
            except Exception as e:
                import traceback
                counts['errors'].append(f"Customer error ({c.get('name', 'unknown')}): {str(e)}\n{traceback.format_exc()}")
    
    # Import Refineries with date ranges
    if 'refineries' in sap_data:
        for r in sap_data['refineries']:
            try:
                plant = get_or_create_plant(r['plant_name'], r.get('location', ''))
                refinery = Refinery.objects.create(
                    simulation=new_master, 
                    plant=plant, 
                    name=r['name'],
                    preference=r.get('preference', 1)
                )
                
                # Handle date_ranges array
                if 'date_ranges' in r:
                    for dr in r['date_ranges']:
                        RefineryDate.objects.create(
                            refinery=refinery,
                            from_date=parse_date(dr['from_date']), 
                            to_date=parse_date(dr['to_date']),
                            daily_refinery_demand=float(dr['daily_refinery_demand'])
                        )
                else:
                    # Legacy format
                    RefineryDate.objects.create(
                        refinery=refinery,
                        from_date=parse_date(r['from_date']), 
                        to_date=parse_date(r['to_date']),
                        daily_refinery_demand=float(r['daily_refinery_demand'])
                    )
                counts['refineries'] += 1
            except Exception as e:
                counts['errors'].append(f"Refinery error ({r.get('name', 'unknown')}): {str(e)}")
    
    # Update last sync time on new master
    new_master.last_sap_sync = datetime.now()
    new_master.save()
    
    print(f"\n{'='*60}")
    print(f"✅ REFRESH COMPLETE")
    print(f"  New Master: {new_master.name} (ID: {new_master.id})")
    print(f"  Version: {version_number}")
    print(f"  Data Source: {'SAP API' if api_used else 'Mock Data'}")
    print(f"  Records Imported:")
    print(f"    - Plants: {counts['plants_created'] + counts['plants_existing']} (New: {counts['plants_created']})")
    print(f"    - Suppliers: {counts['suppliers']}")
    print(f"    - Customers: {counts['customers']}")
    print(f"    - Refineries: {counts['refineries']}")
    print(f"    - Cargos: {counts['cargos']}")
    print(f"    - Plant Inventories: {counts['plant_inventories']}")
    if counts['errors']:
        print(f"  ⚠️ Errors: {len(counts['errors'])}")
    print(f"{'='*60}\n")
    
    # Build success message
    source_text = "SAP API" if api_used else "mock data (SAP API unavailable)"
    success_msg = f'✅ Master data refreshed successfully from {source_text}.\n\n'
    success_msg += f'📦 New Master Version: {version_number}\n'
    success_msg += f'📊 Records Imported:\n'
    success_msg += f'   - {counts["plants_created"] + counts["plants_existing"]} Plants '
    if counts['plants_created'] > 0:
        success_msg += f'(Created: {counts["plants_created"]}, Existing: {counts["plants_existing"]})'
    success_msg += '\n'
    success_msg += f'   - {counts["suppliers"]} Suppliers\n'
    success_msg += f'   - {counts["customers"]} Customers\n'
    success_msg += f'   - {counts["refineries"]} Refineries\n'
    success_msg += f'   - {counts["cargos"]} Cargos\n'
    success_msg += f'   - {counts["plant_inventories"]} Plant Inventories\n'
    success_msg += f'⏰ Sync Time: {new_master.last_sap_sync.strftime("%Y-%m-%d %H:%M:%S")}'
    
    messages.success(request, success_msg)
    
    # Show warnings for any errors
    if counts['errors']:
        for error in counts['errors'][:5]:
            messages.warning(request, f'⚠️ {error}')
        if len(counts['errors']) > 5:
            messages.warning(request, f'⚠️ ... and {len(counts["errors"]) - 5} more errors')
    
    return redirect('lng_planner:dashboard')


def get_mock_sap_data():
    """Return mock SAP data for testing when real API is unavailable."""
    return {
        "plant_inventories": [
            {"plant_name": "Dahej", "opening_inventory": 250},
            {"plant_name": "Ennore", "opening_inventory": 150},
        ],

        "suppliers": [
            {
                "name": "Shell",
                "plant_name": "Dahej",
                "date_ranges": [
                    {"from_date": "2026-01-01", "to_date": "2026-12-31", "daily_supply": 10}
                ]
            },
            {
                "name": "Total",
                "plant_name": "Dahej",
                "date_ranges": [
                    {"from_date": "2026-01-01", "to_date": "2026-12-31", "daily_supply": 8}
                ]
            },
            {
                "name": "QatarGas",
                "plant_name": "Dahej",
            "date_ranges": [
                {"from_date": "2026-02-01", "to_date": "2026-12-31", "daily_supply": 6}
            ]
        },
        {
            "name": "ADNOC LNG",
            "plant_name": "Dahej",
            "date_ranges": [
                {"from_date": "2026-03-01", "to_date": "2026-12-31", "daily_supply": 5}
            ]
        },
        {
            "name": "Petronas",
            "plant_name": "Dahej",
            "date_ranges": [
                {"from_date": "2026-01-15", "to_date": "2026-12-31", "daily_supply": 7}
            ]
        },
        {
            "name": "Chevron",
            "plant_name": "Ennore",
            "date_ranges": [
                {"from_date": "2026-01-01", "to_date": "2026-12-31", "daily_supply": 9}
            ]
        },
        {
            "name": "ExxonMobil",
            "plant_name": "Ennore",
            "date_ranges": [
                {"from_date": "2026-01-01", "to_date": "2026-08-31", "daily_supply": 6}
            ]
        },
        {
            "name": "Cheniere",
            "plant_name": "Ennore",
            "date_ranges": [
                {"from_date": "2026-04-01", "to_date": "2026-12-31", "daily_supply": 5}
            ]
        },
        {
            "name": "Woodside",
            "plant_name": "Ennore",
            "date_ranges": [
                {"from_date": "2026-01-01", "to_date": "2026-06-30", "daily_supply": 4}
            ]
        },
        {
            "name": "Oman LNG",
            "plant_name": "Ennore",
            "date_ranges": [
                {"from_date": "2026-02-01", "to_date": "2026-12-31", "daily_supply": 7}
            ]
        }
    ],

    "customers": [
        {
            "name": "GAIL",
            "plant_name": "Dahej",
            "date_ranges": [
                {"supplier_name": "Shell", "from_date": "2026-01-01", "to_date": "2026-12-31", "daily_demand": 4}
            ]
        },
        {
            "name": "IOCL",
            "plant_name": "Dahej",
            "date_ranges": [
                {"supplier_name": "Total", "from_date": "2026-01-01", "to_date": "2026-12-31", "daily_demand": 3}
            ]
        },
        {
            "name": "BPCL",
            "plant_name": "Dahej",
            "date_ranges": [
                {"supplier_name": "QatarGas", "from_date": "2026-02-01", "to_date": "2026-12-31", "daily_demand": 2}
            ]
        },
        {
            "name": "HPCL",
            "plant_name": "Dahej",
            "date_ranges": [
                {"supplier_name": "ADNOC LNG", "from_date": "2026-03-01", "to_date": "2026-12-31", "daily_demand": 2}
            ]
        },
        {
            "name": "Torrent Gas",
            "plant_name": "Dahej",
            "date_ranges": [
                {"supplier_name": "Petronas", "from_date": "2026-01-15", "to_date": "2026-12-31", "daily_demand": 2}
            ]
        },
        {
            "name": "IGL",
            "plant_name": "Ennore",
            "date_ranges": [
                {"supplier_name": "Chevron", "from_date": "2026-01-01", "to_date": "2026-12-31", "daily_demand": 3}
            ]
        },
        {
            "name": "MGL",
            "plant_name": "Ennore",
            "date_ranges": [
                {"supplier_name": "ExxonMobil", "from_date": "2026-01-01", "to_date": "2026-08-31", "daily_demand": 2}
            ]
        },
        {
            "name": "AG&P",
            "plant_name": "Ennore",
            "date_ranges": [
                {"supplier_name": "Cheniere", "from_date": "2026-04-01", "to_date": "2026-12-31", "daily_demand": 2}
            ]
        },
        {
            "name": "Nayara",
            "plant_name": "Ennore",
            "date_ranges": [
                {"supplier_name": "Woodside", "from_date": "2026-01-01", "to_date": "2026-06-30", "daily_demand": 1}
            ]
        },
        {
            "name": "Adani Total Gas",
            "plant_name": "Ennore",
            "date_ranges": [
                {"supplier_name": "Oman LNG", "from_date": "2026-02-01", "to_date": "2026-12-31", "daily_demand": 3}
            ]
        }
    ],

    "refineries": [
        {
            "name": "IOCL Refinery",
            "plant_name": "Dahej",
            "date_ranges": [
                {"from_date": "2026-01-01", "to_date": "2026-12-31", "daily_refinery_demand": 20}
            ]
        },
        {
            "name": "BPCL Refinery",
            "plant_name": "Dahej",
            "date_ranges": [
                {"from_date": "2026-02-01", "to_date": "2026-12-31", "daily_refinery_demand": 10}
            ]
        },
        {
            "name": "CPCL Refinery",
            "plant_name": "Ennore",
            "date_ranges": [
                {"from_date": "2026-01-01", "to_date": "2026-12-31", "daily_refinery_demand": 15}
            ]
        }
    ],

    "cargos": [
        {"cargo_name": "CARGO-001", "plant_name": "Dahej", "delivery_date": "2026-02-15", "amount": 50},
        {"cargo_name": "CARGO-002", "plant_name": "Dahej", "delivery_date": "2026-03-15", "amount": 40},
        {"cargo_name": "CARGO-003", "plant_name": "Dahej", "delivery_date": "2026-05-10", "amount": 60},
        {"cargo_name": "CARGO-004", "plant_name": "Dahej", "delivery_date": "2026-07-20", "amount": 45},
        {"cargo_name": "CARGO-005", "plant_name": "Dahej", "delivery_date": "2026-10-01", "amount": 55},
        {"cargo_name": "CARGO-006", "plant_name": "Ennore", "delivery_date": "2026-02-20", "amount": 30},
        {"cargo_name": "CARGO-007", "plant_name": "Ennore", "delivery_date": "2026-04-10", "amount": 35},
        {"cargo_name": "CARGO-008", "plant_name": "Ennore", "delivery_date": "2026-06-15", "amount": 40},
        {"cargo_name": "CARGO-009", "plant_name": "Ennore", "delivery_date": "2026-08-25", "amount": 25},
        {"cargo_name": "CARGO-010", "plant_name": "Ennore", "delivery_date": "2026-11-10", "amount": 50}
    ]
    }
